from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from datetime import date

wb = Workbook()

# ─── Color palette ───────────────────────────────────────────────────────────
DARK_GREEN  = "1B5E20"
MID_GREEN   = "2E7D32"
LIGHT_GREEN = "C8E6C9"
ACCENT_GREEN= "4CAF50"
HEADER_GRAY = "37474F"
ROW_ALT     = "F1F8E9"
WHITE       = "FFFFFF"
GOLD        = "F9A825"
LIGHT_GOLD  = "FFF9C4"
BLUE_DARK   = "0D47A1"
BLUE_LIGHT  = "BBDEFB"
RED_DARK    = "B71C1C"
RED_LIGHT   = "FFCDD2"
ORANGE      = "E65100"
ORANGE_LIGHT= "FFE0B2"
PURPLE      = "4A148C"
PURPLE_LIGHT= "E1BEE7"
TEAL        = "004D40"
TEAL_LIGHT  = "B2DFDB"
SECTION_COLORS = {
    "FR-01": ("1B5E20", "C8E6C9"),  # Citizen Reg & Engagement
    "FR-02": ("0D47A1", "BBDEFB"),  # LGA Registration
    "FR-03": ("E65100", "FFE0B2"),  # Subscription & Tenure
    "FR-04": ("4A148C", "E1BEE7"),  # Allocation Tracking
    "FR-05": ("004D40", "B2DFDB"),  # Interactive Map
    "FR-06": ("BF360C", "FFCCBC"),  # LGA Project Showcase
    "FR-07": ("1A237E", "C5CAE9"),  # Public Engagement
    "FR-08": ("33691E", "DCEDC8"),  # Advertising & Ads
    "FR-09": ("880E4F", "F8BBD9"),  # Payment & Notifications
    "FR-10": ("212121", "F5F5F5"),  # Admin Panel
    "FR-11": ("006064", "B2EBF2"),  # Chairman Dashboard
    "FR-12": ("4E342E", "D7CCC8"),  # Data Management
    "FR-13": ("1B5E20", "F9FBE7"),  # Government Collaboration
    "FR-14": ("37474F", "ECEFF1"),  # News & Press
    "FR-15": ("6A1B9A", "F3E5F5"),  # Additional Features
}

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="212121"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def center_align(v="center"):
    return Alignment(horizontal="center", vertical=v, wrap_text=True)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ─── SHEET 1: COVER PAGE ─────────────────────────────────────────────────────
ws_cover = wb.active
ws_cover.title = "Cover Page"
ws_cover.sheet_view.showGridLines = False

ws_cover.merge_cells("A1:J3")
ws_cover["A1"].value = ""
ws_cover["A1"].fill = fill(DARK_GREEN)

ws_cover.merge_cells("A4:J6")
ws_cover["A4"].value = "FEATURES REQUIREMENT DOCUMENT (FRD)"
ws_cover["A4"].font = Font(name="Arial", size=28, bold=True, color=WHITE)
ws_cover["A4"].fill = fill(DARK_GREEN)
ws_cover["A4"].alignment = center_align()

ws_cover.merge_cells("A7:J8")
ws_cover["A7"].value = "Local Government Area (LGA) Citizen Portal"
ws_cover["A7"].font = Font(name="Arial", size=18, bold=True, color=LIGHT_GREEN)
ws_cover["A7"].fill = fill(MID_GREEN)
ws_cover["A7"].alignment = center_align()

ws_cover.merge_cells("A9:J10")
ws_cover["A9"].value = ""
ws_cover["A9"].fill = fill(ACCENT_GREEN)

# Metadata table
meta = [
    ("Project Name", "LGA Citizen Portal – Digital Governance & Engagement Platform"),
    ("Client", "Mr. Abdul Adejoh"),
    ("Contractor", "KVL Systems and Solutions Limited"),
    ("Prepared By", "Mr. Kennedy Egwuda"),
    ("PRD Reference", "LGA_Updated_PRD_v2.pdf – Version 2.0"),
    ("FRD Version", "1.0"),
    ("Date", "May 13, 2026"),
    ("Project Timeline", "4 Months (Phase 1–3) + Ongoing Maintenance (Phase 4)"),
    ("Document Status", "Final"),
    ("Total Features", "62 Functional Requirements + 12 Non-Functional Requirements"),
]

start_row = 12
for i, (label, value) in enumerate(meta):
    r = start_row + i
    ws_cover.merge_cells(f"B{r}:C{r}")
    ws_cover.merge_cells(f"D{r}:J{r}")
    ws_cover[f"B{r}"].value = label
    ws_cover[f"B{r}"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
    ws_cover[f"B{r}"].fill = fill(HEADER_GRAY)
    ws_cover[f"B{r}"].alignment = wrap_align("left", "center")
    ws_cover[f"B{r}"].border = thin_border()
    ws_cover[f"D{r}"].value = value
    ws_cover[f"D{r}"].font = body_font(11)
    ws_cover[f"D{r}"].fill = fill(ROW_ALT if i % 2 == 0 else WHITE)
    ws_cover[f"D{r}"].alignment = wrap_align("left", "center")
    ws_cover[f"D{r}"].border = thin_border()
    ws_cover.row_dimensions[r].height = 22

# Purpose description
r = start_row + len(meta) + 2
ws_cover.merge_cells(f"B{r}:J{r}")
ws_cover[f"B{r}"].value = "DOCUMENT PURPOSE"
ws_cover[f"B{r}"].font = header_font(12, color=WHITE)
ws_cover[f"B{r}"].fill = fill(MID_GREEN)
ws_cover[f"B{r}"].alignment = center_align()
ws_cover[f"B{r}"].border = thin_border()
ws_cover.row_dimensions[r].height = 24

r += 1
ws_cover.merge_cells(f"B{r}:J{r+4}")
desc = (
    "This Features Requirement Document (FRD) provides a comprehensive breakdown of all features to be "
    "developed for the LGA Citizen Portal. It is derived from the Product Requirements Document (PRD v2.0) "
    "and serves as the primary reference for developers, designers, testers, and project managers. Each "
    "feature is catalogued with its unique ID, module classification, description, user roles, acceptance "
    "criteria, development phase, priority, and dependencies. The document covers 15 functional modules "
    "encompassing 62 features and 4 non-functional requirement categories."
)
ws_cover[f"B{r}"].value = desc
ws_cover[f"B{r}"].font = body_font(11)
ws_cover[f"B{r}"].fill = fill(WHITE)
ws_cover[f"B{r}"].alignment = wrap_align("left", "top")
ws_cover[f"B{r}"].border = thin_border()
ws_cover.row_dimensions[r].height = 80

# Module summary table on cover
r = r + 6
ws_cover.merge_cells(f"B{r}:J{r}")
ws_cover[f"B{r}"].value = "MODULE OVERVIEW"
ws_cover[f"B{r}"].font = header_font(12, color=WHITE)
ws_cover[f"B{r}"].fill = fill(DARK_GREEN)
ws_cover[f"B{r}"].alignment = center_align()
ws_cover[f"B{r}"].border = thin_border()
ws_cover.row_dimensions[r].height = 24

r += 1
module_headers = ["Module ID", "Module Name", "Feature Count", "Primary User Roles", "Phase"]
col_letters = ["B", "C", "E", "G", "J"]
col_spans   = [("B","B"),("C","D"),("E","F"),("G","I"),("J","J")]
for (c1,c2), h in zip(col_spans, module_headers):
    ws_cover.merge_cells(f"{c1}{r}:{c2}{r}")
    ws_cover[f"{c1}{r}"].value = h
    ws_cover[f"{c1}{r}"].font = header_font(10, color=WHITE)
    ws_cover[f"{c1}{r}"].fill = fill(HEADER_GRAY)
    ws_cover[f"{c1}{r}"].alignment = center_align()
    ws_cover[f"{c1}{r}"].border = thin_border()
ws_cover.row_dimensions[r].height = 22

modules_overview = [
    ("FR-01","Citizen Registration & Engagement", 5,"Citizens","Phase 1"),
    ("FR-02","LGA Registration & Verification", 5,"LGA / Admin","Phase 1"),
    ("FR-03","Subscription & Tenure Management", 5,"LGA / Admin","Phase 2"),
    ("FR-04","Federal Allocation Tracking", 4,"Citizens / Admin","Phase 2"),
    ("FR-05","Interactive Map & Location-Based Viewing", 5,"All Users","Phase 1"),
    ("FR-06","LGA Project Showcase", 4,"LGA","Phase 1"),
    ("FR-07","Public Engagement & Transparency", 6,"Citizens / Admin","Phase 1"),
    ("FR-08","Advertising & Ads Placement", 6,"Advertisers / Admin","Phase 2"),
    ("FR-09","Payment & Notifications", 4,"LGA / Advertisers / Admin","Phase 2"),
    ("FR-10","Admin Panel & Management", 6,"Admin","Phase 2"),
    ("FR-11","Chairman Dashboard", 6,"LGA Chairman","Phase 1–2"),
    ("FR-12","Data Management & Archiving", 5,"LGA / Admin / Citizens","Phase 2"),
    ("FR-13","Government Collaboration & Transparency", 3,"Admin / Citizens","Phase 2"),
    ("FR-14","News & Press Integration", 3,"Admin / Citizens","Phase 2"),
    ("FR-15","Additional Features & Future Expansion", 5,"All Users","Phase 2–Future"),
]

for i, (mid, mname, fcount, roles, phase) in enumerate(modules_overview):
    r += 1
    row_fill = ROW_ALT if i % 2 == 0 else WHITE
    data = [(("B","B"), mid), (("C","D"), mname), (("E","F"), fcount),
            (("G","I"), roles), (("J","J"), phase)]
    for (c1,c2), val in data:
        ws_cover.merge_cells(f"{c1}{r}:{c2}{r}")
        ws_cover[f"{c1}{r}"].value = val
        ws_cover[f"{c1}{r}"].font = body_font(10)
        ws_cover[f"{c1}{r}"].fill = fill(row_fill)
        ws_cover[f"{c1}{r}"].alignment = center_align() if c1 in ("B","E","J") else wrap_align("left","center")
        ws_cover[f"{c1}{r}"].border = thin_border()
    ws_cover.row_dimensions[r].height = 20

for col in ["A","B","C","D","E","F","G","H","I","J"]:
    ws_cover.column_dimensions[col].width = 14
ws_cover.column_dimensions["A"].width = 3
ws_cover.column_dimensions["C"].width = 22
ws_cover.column_dimensions["G"].width = 22

for i in range(1, 11):
    ws_cover.row_dimensions[i].height = 18

# ─── SHEET 2: FUNCTIONAL REQUIREMENTS ────────────────────────────────────────
ws_fr = wb.create_sheet("Functional Requirements")
ws_fr.sheet_view.showGridLines = False

# Column definitions
fr_headers = [
    "Feature ID", "Module", "Feature Name", "Feature Description",
    "User Role(s)", "Priority", "Acceptance Criteria",
    "Dev Phase", "Milestone", "Dependencies", "Tech Notes", "Status"
]
col_widths = [12, 22, 24, 52, 22, 10, 52, 12, 14, 28, 30, 12]

# Title row
ws_fr.merge_cells("A1:L1")
ws_fr["A1"].value = "FEATURES REQUIREMENT DOCUMENT – FUNCTIONAL REQUIREMENTS"
ws_fr["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_fr["A1"].fill = fill(DARK_GREEN)
ws_fr["A1"].alignment = center_align()
ws_fr.row_dimensions[1].height = 30

ws_fr.merge_cells("A2:L2")
ws_fr["A2"].value = "LGA Citizen Portal | KVL Systems and Solutions Limited | Version 1.0 | May 2026"
ws_fr["A2"].font = Font(name="Arial", size=10, italic=True, color=WHITE)
ws_fr["A2"].fill = fill(MID_GREEN)
ws_fr["A2"].alignment = center_align()
ws_fr.row_dimensions[2].height = 18

# Header row
for col_idx, (h, w) in enumerate(zip(fr_headers, col_widths), start=1):
    cell = ws_fr.cell(row=3, column=col_idx, value=h)
    cell.font = header_font(10, color=WHITE)
    cell.fill = fill(HEADER_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()
    ws_fr.column_dimensions[get_column_letter(col_idx)].width = w
ws_fr.row_dimensions[3].height = 28
ws_fr.freeze_panes = "A4"

# All functional requirements data
features = [
    # ── MODULE FR-01: Citizen Registration & Engagement ────────────────────
    ("FR-01-01","Citizen Registration & Engagement","Citizen Registration",
     "Allow every Nigerian citizen to self-register on the platform for free. Registration requires: full name, email address, phone number, State, and Local Government Area (LGA). Account creation is open to all without restriction.",
     "Citizen","High",
     "1. Registration form collects all required fields (name, email, phone, State, LGA).\n2. Email and phone number are validated for correct format.\n3. Duplicate email/phone detection with appropriate error message.\n4. Account is created and a confirmation email/SMS is sent.\n5. User is redirected to their personalised dashboard on success.",
     "Phase 1","Milestone 1","None",
     "Next.js registration form; PostgreSQL user table; OTP via SMS/email; JWT session on success.",
     "Not Started"),

    ("FR-01-02","Citizen Registration & Engagement","Citizen Dashboard",
     "Upon successful registration, each citizen receives a personalised dashboard showing real-time updates from their selected State and LGA, including: ongoing and completed projects, federal allocation data, and engagement metrics (views, reactions, comments).",
     "Citizen","High",
     "1. Dashboard loads within 4 seconds.\n2. Shows latest posts, projects, and allocation data for citizen's LGA.\n3. Engagement metrics (total reactions, comments, views) are visible.\n4. Citizen can update their State/LGA preference from the dashboard.\n5. Empty-state UI shown when no LGA data is available.",
     "Phase 1","Milestone 1","FR-01-01, FR-02-01",
     "Server-side rendering (SSR) in Next.js; PostgreSQL queries filtered by LGA ID; real-time updates via polling or WebSockets.",
     "Not Started"),

    ("FR-01-03","Citizen Registration & Engagement","Engagement Tools",
     "Citizens can interact with all LGA posts and project updates using: Like, Dislike, Comment, Feedback submission, and Flag/Report for inappropriate content. All interactions are recorded and linked to the citizen's account.",
     "Citizen","High",
     "1. Like/Dislike toggles on every post; counts update in real-time.\n2. Comment form allows text input with max 500 characters; submitted comments appear immediately.\n3. Feedback form allows detailed structured feedback.\n4. Flag/Report button is accessible on each post; flagged posts are queued for admin review.\n5. Citizens cannot interact anonymously (login required).",
     "Phase 1","Milestone 1","FR-01-01, FR-02-01",
     "API routes for reactions (POST /api/reactions); Comments stored in PostgreSQL; WebSocket or polling for live counts.",
     "Not Started"),

    ("FR-01-04","Citizen Registration & Engagement","LGA Citizen Statistics",
     "The system tracks and publicly displays how many registered citizens are associated with each LGA. LGA administrators can view this statistic on their dashboard to gauge citizen interest and platform participation for their area.",
     "Citizen, LGA Chairman","Medium",
     "1. Count of registered citizens per LGA is displayed on the LGA's public page.\n2. LGA Chairman dashboard shows a real-time citizen count widget.\n3. Admin dashboard shows a table of all LGAs with their citizen counts.\n4. Count updates automatically when new citizens register for an LGA.",
     "Phase 1","Milestone 1","FR-01-01, FR-02-01",
     "Aggregate SQL query on users table grouped by LGA ID; cached periodically to avoid expensive re-computation.",
     "Not Started"),

    ("FR-01-05","Citizen Registration & Engagement","OTP Authentication",
     "Multi-factor authentication via One-Time Password (OTP) delivered through SMS and/or email. OTP is required during citizen registration (to verify phone/email), on every login attempt, and for sensitive account actions.",
     "Citizen, LGA Chairman, LGA Staff, Advertiser","High",
     "1. OTP is generated server-side (6-digit numeric code), valid for 5 minutes.\n2. OTP is sent to both email and phone (user selects preferred channel).\n3. Maximum 3 failed OTP attempts before a 15-minute lock.\n4. Resend option available after 60 seconds.\n5. OTP is invalidated after successful use.",
     "Phase 1","Milestone 1","None",
     "SMTP for email OTP; Termii, Africa's Talking, or Twilio for SMS OTP; OTP table in PostgreSQL with expiry timestamp.",
     "Not Started"),

    # ── MODULE FR-02: LGA Registration & Verification ─────────────────────
    ("FR-02-01","LGA Registration & Verification","LGA Profile Creation",
     "Each Local Government must register on the portal by providing a comprehensive profile. Required fields: LGA name, State, Chairman's full name, contact information (phone, email), office address, LGA population data, key sectors (Health, Education, Infrastructure, Agriculture, Water, etc.), and a brief LGA description.",
     "LGA Chairman","High",
     "1. Multi-step registration form collects all mandatory LGA details.\n2. LGA name + State combination must be unique in the system.\n3. Chairman's name and contact info are captured.\n4. At least one sector must be selected.\n5. LGA profile is saved with 'Pending Approval' status after submission.\n6. Chairman receives a confirmation email with reference number.",
     "Phase 1","Milestone 1","FR-01-05",
     "Dedicated LGA table in PostgreSQL; admin approval workflow; pre-seeded list of Nigeria's 774 LGAs and 36 States for validation.",
     "Not Started"),

    ("FR-02-02","LGA Registration & Verification","Verification Process",
     "Before an LGA profile is activated, the Chairman must submit verification documents: (1) Government-issued ID, (2) Unique LGA code provided by the government, (3) OTP authentication via SMS/email, and (4) Certificate of Election for the Chairman.",
     "LGA Chairman, Admin","High",
     "1. Document upload module accepts PDF, JPG, PNG (max 5MB per file).\n2. All four verification items must be submitted before the application is forwarded to admin.\n3. Admin receives a notification when a new verification package is submitted.\n4. LGA profile remains in 'Pending' state until admin approves.\n5. Chairman is notified by email of approval/rejection with reason.",
     "Phase 1","Milestone 1","FR-02-01",
     "File upload to cloud storage (AWS S3 or DigitalOcean Spaces); admin review UI in the admin panel.",
     "Not Started"),

    ("FR-02-03","LGA Registration & Verification","Admin Approval Workflow",
     "All new LGA profiles require manual admin approval before they are activated on the platform. The approval process takes 3–5 business days. Admin reviews submitted profile details and verification documents, then either approves or rejects the application with a reason.",
     "Admin","High",
     "1. Admin panel displays a 'Pending Approvals' queue with all submitted LGA profiles.\n2. Admin can view full LGA profile details and uploaded documents.\n3. Admin can Approve or Reject with a mandatory rejection reason.\n4. Chairman receives an email notification of the outcome.\n5. Approved LGAs are immediately activated and appear on the platform.\n6. Rejected LGAs can resubmit after addressing feedback.",
     "Phase 1","Milestone 1","FR-02-01, FR-02-02",
     "Admin panel built in Next.js; approval status field on LGA table; automated email on status change.",
     "Not Started"),

    ("FR-02-04","LGA Registration & Verification","Admin Charges & Free Period",
     "LGA registration is free for the first 3 months from the platform's public launch date. After this introductory period, an admin fee applies for continued registration and access to the platform. All fees are processed via Paystack.",
     "LGA Chairman, Admin","High",
     "1. New LGAs are flagged with the platform launch date in the database.\n2. System automatically calculates when the 3-month free period ends.\n3. A reminder email is sent 14 days and 7 days before the free period expires.\n4. Paystack payment is triggered at the end of the free period.\n5. LGA access is suspended if payment is not made within the grace period.",
     "Phase 2","Milestone 2","FR-02-01, FR-09-01",
     "Paystack subscription API; cron job to check expiry dates daily; SMTP notification on expiry approach.",
     "Not Started"),

    ("FR-02-05","LGA Registration & Verification","Staff Registration",
     "Each LGA Chairman can register up to two (2) staff members who are granted limited posting permissions. Staff accounts are sub-accounts under the Chairman's profile. Any content (posts, updates) created by staff members requires explicit Chairman approval before being published on the platform.",
     "LGA Chairman, LGA Staff","Medium",
     "1. Chairman can add up to 2 staff accounts via their dashboard.\n2. Staff accounts require name, email, phone, and role.\n3. Staff can draft posts but cannot publish independently.\n4. Chairman receives an in-platform notification and email when a staff member submits a post for review.\n5. Chairman can Approve or Reject staff posts with comments.\n6. Staff posting rights are suspended if Chairman's subscription lapses.",
     "Phase 1","Milestone 1","FR-02-01, FR-03-01",
     "Staff sub-account table linked to LGA/Chairman ID; role-based access control (RBAC); approval workflow similar to admin approval.",
     "Not Started"),

    # ── MODULE FR-03: Subscription & Tenure Management ─────────────────────
    ("FR-03-01","Subscription & Tenure Management","LGA Subscription Fee",
     "After the initial 3-month free period, LGAs must pay a recurring subscription fee via Paystack to maintain active platform access and posting rights. Pricing: Monthly plan = NGN 10,000 | Annual plan = NGN 100,000 (2 months free).",
     "LGA Chairman, Admin","High",
     "1. Two subscription plans are available: Monthly (NGN 10,000) and Annual (NGN 100,000).\n2. Paystack checkout is triggered when selecting a plan.\n3. Successful payment activates the subscription and extends the access period.\n4. Both admin and Chairman receive payment confirmation emails.\n5. Payment appears in Chairman's payment history dashboard.\n6. Annual plan is displayed as 'Save NGN 20,000' compared to monthly.",
     "Phase 2","Milestone 2","FR-02-04, FR-09-01",
     "Paystack Plans API for recurring billing; webhook to handle payment events; plan details stored in LGA subscription table.",
     "Not Started"),

    ("FR-03-02","Subscription & Tenure Management","Tenure Management",
     "LGA posting rights and platform access are tied directly to the Chairman's tenure in office. When a Chairman's tenure officially ends, all posting rights and active subscriptions are automatically suspended. The LGA profile page remains visible but no new content can be published.",
     "LGA Chairman, Admin","High",
     "1. Chairman tenure end date is captured during registration.\n2. System monitors tenure end dates via daily cron job.\n3. 30-day and 7-day reminder emails are sent before tenure expires.\n4. Posting rights are suspended on the tenure end date.\n5. Subscription is paused (not cancelled) pending re-election verification.\n6. LGA public page remains accessible to citizens with 'Tenure Ended' label.",
     "Phase 2","Milestone 2","FR-02-01, FR-03-01",
     "Tenure_end_date field on LGA table; cron job for daily checks; status field: Active / Tenure-Ended / Suspended.",
     "Not Started"),

    ("FR-03-03","Subscription & Tenure Management","Re-election Process",
     "Chairmen seeking to continue using the platform after re-election must formally notify the platform and submit proof of re-election. Required: Certificate of Election. The same admin verification process applies (3–5 business days for admin confirmation).",
     "LGA Chairman, Admin","High",
     "1. Chairman can submit a 'Re-election Notification' from their dashboard.\n2. Certificate of Election (PDF/JPG/PNG) is uploaded as proof.\n3. Admin receives notification and reviews within 3–5 business days.\n4. On approval, Chairman's tenure end date is updated and access is restored.\n5. Chairman receives confirmation email with new tenure details.\n6. Historical data from previous tenure is preserved and linked.",
     "Phase 2","Milestone 2","FR-02-02, FR-02-03, FR-03-02",
     "Re-election workflow mirrors initial verification; new tenure record created in DB linked to existing LGA.",
     "Not Started"),

    ("FR-03-04","Subscription & Tenure Management","Grace Period",
     "A 7-day grace period is provided to LGAs before their accounts are fully suspended due to subscription expiry or tenure end. During this period, the LGA can still view their dashboard and make payment but cannot publish new content.",
     "LGA Chairman","Medium",
     "1. Grace period of exactly 7 calendar days begins on subscription/tenure expiry.\n2. Chairman can still log in and view the dashboard during grace period.\n3. Publishing new content is disabled (greyed out with a notice).\n4. Daily reminder emails are sent during the grace period.\n5. Account is fully suspended after 7 days if no action is taken.\n6. Suspended accounts can be reactivated by making payment or completing re-election.",
     "Phase 2","Milestone 2","FR-03-01, FR-03-02",
     "Grace period logic in cron job; status states: Active → Grace → Suspended; email template for each grace-period day.",
     "Not Started"),

    ("FR-03-05","Subscription & Tenure Management","Staff Posting Rights Control",
     "Staff members' ability to post on the platform is always contingent on their Chairman having an active, paid subscription. If the Chairman's subscription lapses, expires, or is suspended, all associated staff posting rights are automatically revoked.",
     "LGA Staff, LGA Chairman","Medium",
     "1. Staff posting rights are checked against Chairman's subscription status on every post attempt.\n2. If subscription is inactive, the publish/draft button is disabled for staff.\n3. Staff see an informational message explaining why posting is disabled.\n4. Rights are automatically restored when Chairman renews subscription.\n5. This applies to both existing staff posts in draft and new post creation.",
     "Phase 2","Milestone 2","FR-02-05, FR-03-01, FR-03-02",
     "Middleware check on post-creation API route; subscription status lookup from LGA table.",
     "Not Started"),

    # ── MODULE FR-04: Federal Allocation Tracking ──────────────────────────
    ("FR-04-01","Federal Allocation Tracking & Publicising","Monthly Allocation Data",
     "Once the Federal Government releases monthly allocation figures, the portal admin will publish a breakdown of what each of the 36 States and 774 LGAs receives. Citizens can view this data in a clear, accessible, and well-designed format on the platform.",
     "Citizens, Admin","High",
     "1. Allocation data is displayed in a structured table per State and LGA.\n2. Users can filter by State or LGA to find relevant figures.\n3. Data is labelled with the relevant month and year.\n4. Figures are displayed in Nigerian Naira (NGN) with proper formatting.\n5. Admin can publish allocation data via the admin content management panel.\n6. Page is publicly accessible without login.",
     "Phase 2","Milestone 2","FR-10-04",
     "Allocation data table in PostgreSQL (LGA_id, month, year, amount); admin CMS for data entry; optional Excel/CSV import for bulk upload.",
     "Not Started"),

    ("FR-04-02","Federal Allocation Tracking & Publicising","Admin Content Creation for Allocations",
     "The portal administrator is responsible for creating and publishing all content related to federal allocation information. This includes written analysis articles, infographic uploads, and downloadable summary reports for each allocation cycle.",
     "Admin","High",
     "1. Admin CMS allows creation of allocation articles with a rich text editor.\n2. Admin can attach infographic images (JPG, PNG, max 10MB).\n3. Admin can attach PDF summary reports.\n4. Published content appears on the Allocation Tracking section of the platform.\n5. Drafts can be saved before publishing.\n6. Admin can schedule publication for a specific date/time.",
     "Phase 2","Milestone 2","FR-10-04",
     "Rich text editor (TipTap or Quill.js) in admin panel; file upload to CDN; allocation_articles table in PostgreSQL.",
     "Not Started"),

    ("FR-04-03","Federal Allocation Tracking & Publicising","Historical Allocation Archive",
     "All historical allocation data published on the platform is archived and remains publicly searchable. Citizens can search and filter allocation records by: specific month/year, State, and LGA. This enables long-term trend tracking and civic accountability.",
     "Citizens","High",
     "1. Search/filter UI allows selection of month, year, State, and LGA.\n2. Results display in a paginated table (25 rows per page).\n3. All allocation records from platform launch are retained indefinitely.\n4. Users can download allocation data as CSV or PDF.\n5. Comparison of current vs. historical figures is visually indicated.",
     "Phase 2","Milestone 2","FR-04-01",
     "Indexed PostgreSQL queries for allocation archive; PDF/CSV export using server-side generation; pagination API.",
     "Not Started"),

    ("FR-04-04","Federal Allocation Tracking & Publicising","Allocation Comparison Tool",
     "Citizens can use an interactive comparison tool to compare federal allocation amounts across multiple LGAs and/or States side by side. The tool promotes informed civic discourse and accountability by enabling data-driven comparisons.",
     "Citizens","Medium",
     "1. User can select up to 5 LGAs or States to compare simultaneously.\n2. Comparison is displayed as a bar chart and a side-by-side data table.\n3. Users can select a specific time period (month/year range) for comparison.\n4. Charts are rendered client-side and are shareable via link.\n5. Export as image (PNG) or PDF is available.",
     "Phase 2","Milestone 2","FR-04-01, FR-04-03",
     "Chart.js or Recharts for visualisation; shareable URL with query params encoding selected LGAs/period.",
     "Not Started"),

    # ── MODULE FR-05: Interactive Map ──────────────────────────────────────
    ("FR-05-01","Interactive Map & Location-Based Viewing","Interactive Nigeria Map",
     "An interactive map of Nigeria is displayed prominently on the platform, showing all 774 LGAs with visual markers. The map is the primary discovery interface for citizens to find and explore LGA activities across the country.",
     "All Users","High",
     "1. Map renders with all 774 Nigerian LGA boundaries or markers.\n2. Map loads within 3 seconds on a standard 4G connection.\n3. Users can zoom in/out and pan across the map.\n4. Search bar on the map allows users to find a specific LGA by name.\n5. Map is fully responsive on desktop, tablet, and mobile.",
     "Phase 1","Milestone 1","None",
     "Leaflet.js or Mapbox GL JS; GeoJSON data for Nigerian LGA boundaries; PostGIS for spatial queries; tile caching for performance.",
     "Not Started"),

    ("FR-05-02","Interactive Map & Location-Based Viewing","LGA Detail Pop-up on Click",
     "When a user clicks on any LGA marker or boundary on the map, a detail pop-up or side panel is displayed showing: the Chairman's name, latest project posts (with photos/videos thumbnails), most recent public comments, and the latest federal allocation figure for that LGA.",
     "All Users","High",
     "1. Clicking an LGA opens a pop-up/side panel within 1 second.\n2. Pop-up shows: LGA name, Chairman's name, latest 3 project posts, latest allocation figure, comment count.\n3. 'View Full Profile' link navigates to the full LGA page.\n4. Pop-up closes when user clicks elsewhere or on a close button.\n5. For unregistered LGAs, a 'Not Yet Registered' message is shown.",
     "Phase 1","Milestone 1","FR-05-01, FR-02-01",
     "Leaflet.js popup component; API route to fetch LGA summary data; lazy-load images in popup for performance.",
     "Not Started"),

    ("FR-05-03","Interactive Map & Location-Based Viewing","Colour-Coded Project Markers",
     "Project markers on the interactive map use a colour-coding system to visually indicate the project category. Categories and their colours are: Roads/Infrastructure (Orange), Health (Red), Education (Blue), Water (Teal/Cyan), Agriculture (Green). A map legend is displayed to guide users.",
     "All Users","Medium",
     "1. Each active project shows a colour-coded pin on the map at its location.\n2. Map legend is always visible explaining colour-to-category mapping.\n3. Users can filter map markers by project category via a filter panel.\n4. Multiple projects in same location are clustered into a single marker with a count.\n5. Clicking a project marker shows project name, status, and a link to the full project page.",
     "Phase 1","Milestone 1","FR-05-01, FR-06-01",
     "Leaflet.js marker clustering plugin; custom icon colours per category; project geolocation stored in PostGIS.",
     "Not Started"),

    ("FR-05-04","Interactive Map & Location-Based Viewing","Real-Time Project Status Tracking",
     "Citizens can track the status of any project on the map in real-time. Each project has one of three statuses: Completed (Green badge), In Progress (Yellow badge), or Pending (Grey badge). Status badges are visible on map markers and project listings.",
     "All Users","High",
     "1. Project status badge is displayed on the map marker and project card.\n2. Status can be updated by the LGA Chairman or authorised staff.\n3. Status change is logged with a timestamp and the name of who changed it.\n4. Citizens are not able to change project status.\n5. Status changes trigger a notification to citizens who have 'followed' the project.",
     "Phase 1","Milestone 1","FR-05-01, FR-06-01",
     "project_status ENUM field in PostgreSQL; audit log table for status changes; optional push/email notification on status update.",
     "Not Started"),

    ("FR-05-05","Interactive Map & Location-Based Viewing","Archived Projects on Map",
     "Completed and archived projects remain visible on the interactive map and in project listings even after the project is marked as completed or after a Chairman's tenure ends. This ensures permanent public transparency and historical record-keeping.",
     "All Users","Medium",
     "1. Archived projects display a 'Completed' or 'Archived' badge on the map.\n2. Archived projects are included in map results by default but can be filtered out.\n3. Clicking an archived project marker shows full project details including the originating Chairman's tenure.\n4. Archived projects cannot be edited by current Chairman.\n5. Admin can hide specific archived projects if flagged for policy violations.",
     "Phase 1","Milestone 1","FR-05-01, FR-12-01",
     "is_archived flag on projects table; archived projects fetched in same API but with visual differentiation.",
     "Not Started"),

    # ── MODULE FR-06: LGA Project Showcase ────────────────────────────────
    ("FR-06-01","LGA Project Showcase","Project Publication",
     "Registered and verified LGAs can publish details of their completed and ongoing government projects on the platform. Projects can include multimedia content: videos, images, and rich text descriptions. Each project is publicly visible to all platform users.",
     "LGA Chairman, LGA Staff","High",
     "1. Project creation form is available only to verified and active LGAs.\n2. Chairman and authorised staff can create and submit projects.\n3. Published projects appear on the LGA profile page, the interactive map, and in the news feed.\n4. Projects created by staff require Chairman approval before publication.\n5. Published projects are immediately indexed for search.",
     "Phase 1","Milestone 1","FR-02-01, FR-02-03",
     "Projects table in PostgreSQL; post creation API; media upload to CDN; ElasticSearch or PostgreSQL full-text search.",
     "Not Started"),

    ("FR-06-02","LGA Project Showcase","Project Detail Requirements",
     "Each project listing published on the platform must contain the following mandatory and optional fields. Mandatory: project title, description, category, status, location (geolocation pin), budget, start date, expected completion date, supporting media (min 1 photo). Optional: YouTube/video link, additional photos.",
     "LGA Chairman, LGA Staff","High",
     "1. Project form enforces mandatory field completion before submission.\n2. Project title max 150 characters; description max 2000 characters.\n3. Budget field accepts NGN values only, formatted as currency.\n4. Date fields validate that start date is before expected completion date.\n5. At least one image must be uploaded.\n6. Geolocation pin can be set by clicking on an embedded map or entering coordinates.",
     "Phase 1","Milestone 1","FR-06-01",
     "Form validation both client-side (React Hook Form) and server-side; PostGIS point for location; budget stored as INTEGER in kobo (smallest NGN unit).",
     "Not Started"),

    ("FR-06-03","LGA Project Showcase","Verification Prerequisite for Publishing",
     "Only LGAs that have completed the full registration and verification process, received admin approval, and have an active subscription (post-free period) can publish project updates. Unverified or inactive LGAs are blocked from publishing.",
     "LGA Chairman","High",
     "1. System checks LGA verification status and subscription status before allowing project creation.\n2. Unverified LGAs see a locked/disabled 'Create Project' button with explanation.\n3. Suspended LGAs cannot create, edit, or publish projects.\n4. API endpoint for project creation returns 403 Forbidden for non-active LGAs.\n5. Admin can override and grant temporary posting access if needed.",
     "Phase 1","Milestone 1","FR-02-03, FR-03-01",
     "Middleware guard on project creation API checking LGA status; status states enforced at both UI and API layer.",
     "Not Started"),

    ("FR-06-04","LGA Project Showcase","Rich Media Support",
     "Project posts support rich multimedia content including: multiple image uploads (gallery view), direct video file uploads, embedded YouTube or Vimeo video links, and rich text descriptions with formatting (bold, italic, lists, headings). All media is served via CDN.",
     "LGA Chairman, LGA Staff","High",
     "1. Image upload accepts JPG, PNG, WebP; max 10MB per image; max 10 images per project.\n2. Video upload accepts MP4, MOV; max 200MB; progress bar shown during upload.\n3. YouTube/Vimeo URLs are embedded automatically with a preview thumbnail.\n4. Rich text editor supports: bold, italic, underline, ordered/unordered lists, headings (H2/H3).\n5. Uploaded media is stored on CDN (AWS CloudFront or DigitalOcean Spaces CDN).\n6. Images are auto-compressed/resized on upload.",
     "Phase 1","Milestone 1","FR-06-01",
     "Multer or Next.js API route for file upload; Sharp.js for image compression; CDN for delivery; Tiptap for rich text editor.",
     "Not Started"),

    # ── MODULE FR-07: Public Engagement & Transparency ─────────────────────
    ("FR-07-01","Public Engagement & Transparency","Reaction System",
     "Citizens can react to any LGA post or project update using one of five reaction types: Like, Dislike, Support, Question, and Report. Reactions are attributed to logged-in citizens and count totals are displayed publicly on each post.",
     "Citizen","High",
     "1. Five reaction options are displayed on every post and project card.\n2. Citizen can select only one reaction per post (changes are permitted).\n3. Reaction counts are displayed in real-time (or near real-time).\n4. 'Report' reaction flags the post for admin review (also logged separately).\n5. Citizens must be logged in to react; unauthenticated users see counts only.\n6. Total reaction breakdown is visible to LGA Chairman on their dashboard.",
     "Phase 1","Milestone 1","FR-01-01",
     "Reactions table with (user_id, post_id, reaction_type); upsert logic to allow reaction changes; real-time count via WebSockets or SWR polling.",
     "Not Started"),

    ("FR-07-02","Public Engagement & Transparency","Comments & Feedback",
     "Citizens can leave detailed comments and structured feedback on any LGA post or project update. Comments are publicly visible. The LGA Chairman can reply to comments directly from their dashboard. All comment activity is tracked for analytics.",
     "Citizen, LGA Chairman","High",
     "1. Comment input box is available below every post.\n2. Comments are paginated (10 per page); newest first.\n3. Citizens can edit or delete their own comments within 24 hours.\n4. Chairman can reply to comments, and replies are visually nested.\n5. Comment character limit: 500 characters.\n6. Profanity/spam filtered by AI moderation before display.",
     "Phase 1","Milestone 1","FR-01-01, FR-07-05",
     "Comments table (post_id, user_id, content, parent_id for replies); AI moderation via OpenAI Moderation API or similar.",
     "Not Started"),

    ("FR-07-03","Public Engagement & Transparency","Social Media Sharing",
     "Citizens and visitors can share any LGA post, project update, or allocation data via social media directly from the platform. Supported platforms: WhatsApp, Twitter/X, and Facebook. Sharing generates a preview card with the post title, image, and URL.",
     "All Users","Medium",
     "1. Share buttons (WhatsApp, Twitter/X, Facebook) are visible on all posts and project cards.\n2. Each post has a unique, shareable URL.\n3. Shared links generate an OG (Open Graph) preview card with title, description, and image.\n4. WhatsApp share pre-fills the message with post title and URL.\n5. Share buttons work without requiring the user to be logged in.\n6. Share counts are tracked and displayed on each post.",
     "Phase 1","Milestone 1","None",
     "Open Graph meta tags on dynamic post pages; Next.js dynamic OG image generation; Web Share API for mobile; platform-specific share URLs.",
     "Not Started"),

    ("FR-07-04","Public Engagement & Transparency","Content Reporting System",
     "Any citizen can report a post, comment, or project update that they find inappropriate, misleading, or in violation of platform guidelines. Reported content is placed in a moderation queue for admin review. The reporter is notified of the outcome.",
     "Citizen, Admin","High",
     "1. A 'Report' button is accessible on all posts, comments, and project updates.\n2. Reporting user selects a reason from a predefined list (Misleading, Inappropriate, Spam, Offensive, Other).\n3. Optional text box allows additional details (max 200 characters).\n4. Report is logged and content is flagged in the admin moderation queue.\n5. Reporter receives an email acknowledgement.\n6. If content receives 5+ reports, it is automatically hidden pending review.",
     "Phase 1","Milestone 1","FR-01-01, FR-10-06",
     "Reports table (content_id, content_type, reporter_id, reason, details); auto-hide trigger after threshold; admin moderation queue UI.",
     "Not Started"),

    ("FR-07-05","Public Engagement & Transparency","AI-Powered Content Moderation",
     "The platform uses AI-powered content moderation to automatically filter and flag inappropriate language, hate speech, and spam from comments and feedback submissions before they are publicly displayed. The AI operates as a pre-publication filter.",
     "System (Automated)","High",
     "1. All comments and feedback pass through the AI moderation layer before being saved/displayed.\n2. Content flagged as inappropriate is held in a moderation queue (not displayed immediately).\n3. Content scoring above the violation threshold is auto-rejected with an error message to the user.\n4. Admin can review and override AI moderation decisions.\n5. False positive rate must be below 5%; system must process moderation within 2 seconds.",
     "Phase 2","Milestone 2","FR-07-02",
     "OpenAI Moderation API or Google Perspective API integration; moderation_log table; configurable threshold in admin settings.",
     "Not Started"),

    ("FR-07-06","Public Engagement & Transparency","Admin Moderation Authority",
     "Platform administrators have full authority to moderate all user-generated content. Admins can delete any post, comment, or project update that violates platform guidelines. Admins can also temporarily suspend or permanently ban any user account.",
     "Admin","High",
     "1. Admin panel has a dedicated 'Moderation' section with flagged content queue.\n2. Admin can view full context of any post before taking action.\n3. Admin can Delete Content, Warn User, Suspend Account (specify duration), or Ban Account.\n4. All moderation actions are logged with admin ID, timestamp, and reason.\n5. Banned users receive an email notification with the reason.\n6. Banned users cannot re-register with the same email or phone number.",
     "Phase 1","Milestone 1","FR-10-02",
     "Admin role guard on moderation API routes; moderation_actions audit table; ban list checking on registration.",
     "Not Started"),

    # ── MODULE FR-08: Advertising & Ads Placement ──────────────────────────
    ("FR-08-01","Advertising & Ads Placement","Ads Placement System",
     "The platform features a dedicated advertising system enabling businesses, organisations, and individuals to place advertisements on the portal. Advertisements are shown to platform users across various sections of the site, generating revenue for the platform.",
     "Advertiser, Admin","High",
     "1. Advertisers can self-register as advertiser accounts on the platform.\n2. Advertisers can browse available ad plans and slots.\n3. Ad slots are available in defined locations across the platform.\n4. The ads system is distinct from LGA subscription fees.\n5. Admin controls all available ad slots and their visibility.",
     "Phase 2","Milestone 2","FR-08-02, FR-09-01",
     "Advertiser account type in users table; ad_slots table; ad placement logic integrated into Next.js page layouts.",
     "Not Started"),

    ("FR-08-02","Advertising & Ads Placement","Ad Subscription Plans",
     "Advertisers choose from three tiered subscription plans: Basic, Standard, and Premium. Each plan offers different placement locations, ad format options, display duration, and pricing. Plan pricing is set and managed by the admin.",
     "Advertiser","High",
     "1. Three plans are available: Basic, Standard, Premium.\n2. Each plan clearly states: placement locations, ad formats, duration, and price.\n3. Plan comparison table is displayed to the advertiser during checkout.\n4. Advertiser selects a plan and proceeds to Paystack checkout.\n5. Plan is activated immediately upon successful payment.\n6. Admin can update plan details and pricing from the admin panel.",
     "Phase 2","Milestone 2","FR-08-01, FR-09-01",
     "ad_plans table in PostgreSQL; Paystack plans or one-time charges; plan-to-slot mapping table.",
     "Not Started"),

    ("FR-08-03","Advertising & Ads Placement","Ad Formats",
     "The platform supports multiple ad formats to give advertisers flexibility in how they reach citizens. Supported formats: (1) Banner Ads – horizontal display at top/bottom of pages, (2) Sidebar Ads – vertical ads on side panels, (3) Sponsored Posts – integrated in the news feed, (4) Featured Listings – highlighted LGA or project listings.",
     "Advertiser, Admin","Medium",
     "1. Each ad format has defined dimensions (e.g., banner: 728x90px; sidebar: 300x250px).\n2. Advertiser uploads creative assets in the correct format/dimensions.\n3. System validates uploaded creative dimensions and file size (max 2MB per creative).\n4. Sponsored posts appear in the feed with a 'Sponsored' label.\n5. Featured listings appear at the top of LGA search results with a highlighted border.\n6. Ad formats are linked to specific subscription plan tiers.",
     "Phase 2","Milestone 2","FR-08-02",
     "Ad creative upload to CDN; ad delivery integrated into Next.js layout components; responsive ad units.",
     "Not Started"),

    ("FR-08-04","Advertising & Ads Placement","Ad Payment via Paystack",
     "All advertising subscription payments are processed exclusively through Paystack. Upon successful payment, automated email notifications are dispatched to both the platform admin and the advertiser confirming the transaction and activating the ad subscription.",
     "Advertiser, Admin","High",
     "1. Paystack checkout is triggered when advertiser selects a plan.\n2. All supported Paystack payment methods are available (card, bank transfer, USSD).\n3. On payment success (via webhook), ad subscription is immediately activated.\n4. Admin receives email: advertiser name, plan selected, amount, transaction reference.\n5. Advertiser receives email: plan details, activation confirmation, invoice attachment.\n6. Failed payments are logged and the advertiser is notified to retry.",
     "Phase 2","Milestone 2","FR-08-02, FR-09-01",
     "Paystack webhook endpoint; idempotent payment processing; transaction_log table; SMTP email with PDF invoice attachment.",
     "Not Started"),

    ("FR-08-05","Advertising & Ads Placement","Advertiser Self-Service Dashboard",
     "Advertisers have access to a dedicated self-service dashboard where they can manage all aspects of their ad campaigns. The dashboard provides real-time campaign performance data including: impressions, clicks, click-through rate (CTR), and ad spend.",
     "Advertiser","Medium",
     "1. Dashboard shows active, paused, and expired campaigns.\n2. Impression and click counts are updated daily.\n3. Advertiser can pause, resume, or cancel active campaigns.\n4. Advertiser can upload updated ad creatives for active campaigns.\n5. Campaign performance graphs show trends over the subscription period.\n6. Advertiser can download a performance report as PDF or CSV.",
     "Phase 2","Milestone 2","FR-08-01, FR-08-04",
     "Ad impressions tracked via server-side logging; ad_clicks table; dashboard charts using Recharts or Chart.js.",
     "Not Started"),

    ("FR-08-06","Advertising & Ads Placement","Admin Ad Management",
     "The admin has complete, centralised control over the entire advertising system. Admin functions include: creating and managing ad plans, reviewing and approving/rejecting submitted ads for quality, managing available ad slots, setting and updating pricing, and monitoring total advertising revenue.",
     "Admin","High",
     "1. Admin panel has a dedicated 'Ads Management' section.\n2. Admin can create, edit, and deactivate ad subscription plans.\n3. New ad submissions from advertisers appear in an 'Ads Pending Review' queue.\n4. Admin can Approve or Reject ads with a rejection reason sent to the advertiser.\n5. Admin can view a revenue dashboard showing total ad revenue by day, week, and month.\n6. Admin can manage available ad slots (activate/deactivate specific positions on the site).",
     "Phase 2","Milestone 2","FR-08-01, FR-08-02",
     "Admin role guard; ad_review workflow similar to LGA approval; revenue aggregation queries; slot management table.",
     "Not Started"),

    # ── MODULE FR-09: Payment & Notifications ─────────────────────────────
    ("FR-09-01","Payment & Notifications","Paystack Payment Gateway",
     "All financial transactions on the platform are processed exclusively through Paystack. This includes: LGA subscription fees (monthly/annual), admin charges, and ad placement subscription payments. Paystack handles all payment security and PCI-DSS compliance.",
     "LGA Chairman, Advertiser, Admin","High",
     "1. Paystack is the sole payment gateway (no other providers).\n2. Paystack checkout modal is embedded in the platform (not a redirect).\n3. Paystack webhooks are configured to handle: successful payment, failed payment, subscription renewal, and cancellation events.\n4. All Paystack transactions are logged in the platform's transaction_log table.\n5. Platform complies with Paystack's PCI-DSS requirements by not storing card data.",
     "Phase 2","Milestone 2","None",
     "Paystack Node.js SDK; webhook endpoint with signature verification; test mode for development; live mode for production.",
     "Not Started"),

    ("FR-09-02","Payment & Notifications","Automated Email Notifications",
     "Every successful payment on the platform triggers an automated transactional email notification to both the platform admin and the user who made the payment. Notification emails contain: payment amount, date and time, purpose of payment, and unique transaction reference number.",
     "LGA Chairman, Advertiser, Admin","High",
     "1. Email is triggered immediately upon Paystack webhook confirming successful payment.\n2. Admin receives email with full transaction details.\n3. Payer receives email with: amount, date, purpose, reference, and PDF invoice.\n4. Emails use a professionally designed HTML template with platform branding.\n5. Failed emails are retried up to 3 times with 5-minute intervals.\n6. Notification logs are stored in the database for audit.",
     "Phase 2","Milestone 2","FR-09-01",
     "Nodemailer or SendGrid for SMTP/transactional email; Handlebars or React Email for HTML templates; email_notifications_log table.",
     "Not Started"),

    ("FR-09-03","Payment & Notifications","Payment History",
     "All users (citizens with paid services, LGAs, and advertisers) can view their complete and unalterable payment history directly on their respective dashboards. The history shows all transactions made through the platform, filterable by date range and payment type.",
     "LGA Chairman, Advertiser","Medium",
     "1. Payment history is accessible from the user/LGA/advertiser dashboard.\n2. Each entry shows: date, amount, purpose, status (Successful/Failed/Pending), reference.\n3. History is paginated (20 entries per page); newest first.\n4. Users can filter history by date range and payment type.\n5. Each payment entry has a 'Download Invoice' button.\n6. Payment history cannot be edited or deleted by the user.",
     "Phase 2","Milestone 2","FR-09-01",
     "transaction_log table with user_id, lga_id, amount, purpose, status, paystack_reference, created_at.",
     "Not Started"),

    ("FR-09-04","Payment & Notifications","Automatic Invoice Generation",
     "The platform automatically generates a professional PDF invoice for every completed financial transaction. Invoices are attached to payment confirmation emails and are also downloadable from the payment history section of the user's dashboard at any time.",
     "LGA Chairman, Advertiser","Medium",
     "1. Invoice is generated automatically upon payment confirmation.\n2. Invoice includes: invoice number, date, payer details, platform details, itemised charges, total amount, and transaction reference.\n3. Invoice is in PDF format with platform branding and logo.\n4. Invoice is stored on CDN and link is saved in the transaction record.\n5. Invoice is downloadable from payment history.\n6. Invoice numbering follows a sequential, unique format (e.g., INV-2026-000001).",
     "Phase 2","Milestone 2","FR-09-01, FR-09-02",
     "PDFKit or Puppeteer for server-side PDF generation; invoice stored on CDN; invoice_number auto-increment sequence in PostgreSQL.",
     "Not Started"),

    # ── MODULE FR-10: Admin Panel & Management ─────────────────────────────
    ("FR-10-01","Admin Panel & Management","Full LGA Management",
     "The admin has complete oversight and control over all LGA profiles on the platform. Admin actions include: viewing all LGA profiles (with full details), approving or rejecting new LGA applications, suspending active LGAs for policy violations, and deactivating LGAs that have permanently closed or failed compliance. Admin can also view the exact number of registered citizens per LGA.",
     "Admin","High",
     "1. Admin LGA management page lists all 774 LGAs with status, chairman, and citizen count.\n2. Admin can search/filter by State, status, and registration date.\n3. Admin can click any LGA to view full profile details and documents.\n4. Admin can Approve, Reject, Suspend, or Deactivate an LGA with a required reason.\n5. LGA Chairman receives email notification of any status change.\n6. Admin can view real-time citizen registration count per LGA.",
     "Phase 2","Milestone 2","FR-02-01, FR-02-02, FR-02-03",
     "Admin-only API routes with RBAC middleware; LGA management table view; bulk operations support.",
     "Not Started"),

    ("FR-10-02","Admin Panel & Management","User Account Management",
     "Admins can manage all user accounts across all user types: Citizens, LGA Chairmen, LGA Staff, and Advertisers. Admin can view account details, activity history, suspend accounts temporarily (specify duration), or permanently ban accounts for policy violations.",
     "Admin","High",
     "1. Admin user management page lists all accounts with type, registration date, and status.\n2. Admin can search by name, email, phone, or LGA.\n3. Admin can view a user's full activity history (posts, comments, reactions, payments).\n4. Admin can Suspend (specify duration) or Ban (permanent) any account.\n5. Suspended/banned users receive email notification with reason.\n6. Admin can reinstate a suspended or banned account.",
     "Phase 2","Milestone 2","FR-01-01, FR-02-01",
     "Admin user list API; user_status ENUM (Active/Suspended/Banned); ban_list table; audit_log for admin actions.",
     "Not Started"),

    ("FR-10-03","Admin Panel & Management","Ads Placement Management",
     "The admin has complete centralised control over the advertising system. This includes: creating, editing, and managing ad subscription plans; reviewing and approving/rejecting submitted ad creatives; managing the configuration of available ad slots across the platform; updating pricing; and monitoring total advertising revenue analytics.",
     "Admin","High",
     "1. Admin can create new ad plans with custom: name, price, duration, formats, and placement locations.\n2. Admin can activate/deactivate ad plans from the admin panel.\n3. Admin ad review queue shows submitted ads awaiting approval with creative preview.\n4. Admin can approve or reject ads; rejected advertisers receive reason via email.\n5. Revenue dashboard shows: total ad revenue, active campaigns, upcoming renewals.\n6. Admin can configure site-wide ad slot visibility (show/hide any slot globally).",
     "Phase 2","Milestone 2","FR-08-06",
     "Ad plan CRUD API; ad review workflow; ad_slots configuration table; revenue aggregation queries with date filters.",
     "Not Started"),

    ("FR-10-04","Admin Panel & Management","Content Creation & Publishing",
     "Administrators are responsible for creating and publishing all editorial content on the platform, including: federal allocation data and analysis articles, LGA news and updates, platform-wide announcements, infographic uploads, and downloadable reports. A rich text editor with media upload support is provided in the admin panel.",
     "Admin","High",
     "1. Admin CMS panel provides a rich text editor (Tiptap or similar) for content creation.\n2. Admin can create, save as draft, preview, schedule, and publish content.\n3. Content types: Allocation Reports, LGA News, Platform Announcements.\n4. Admin can attach images (JPG/PNG, max 10MB) and PDF files to content.\n5. Published content appears on the public-facing sections immediately.\n6. Admin can edit or unpublish/archive existing content at any time.",
     "Phase 2","Milestone 2","None",
     "Admin CMS with Tiptap rich text editor; content table in PostgreSQL with type, status, scheduled_at fields; CDN for media attachments.",
     "Not Started"),

    ("FR-10-05","Admin Panel & Management","Engagement Analytics Dashboard",
     "The admin has access to a comprehensive analytics dashboard showing platform-wide engagement metrics: total views, comments, and reactions per post/LGA/period; sentiment analysis overview; most active LGAs; top-engaging content; and downloadable monthly analytics reports in PDF format.",
     "Admin","High",
     "1. Admin analytics dashboard displays: total platform users, active LGAs, total posts, total reactions/comments.\n2. Time-series graphs show daily/weekly/monthly engagement trends.\n3. Top 10 most active LGAs by engagement are listed.\n4. Sentiment analysis summary (positive/neutral/negative breakdown) is displayed.\n5. Admin can generate and download a full monthly report as PDF.\n6. All metrics are filterable by date range, State, and LGA.",
     "Phase 2","Milestone 2","FR-11-02",
     "Analytics aggregation queries; Chart.js/Recharts for graphs; Puppeteer or PDFKit for PDF report generation; monthly cron job for report generation.",
     "Not Started"),

    ("FR-10-06","Admin Panel & Management","Content Moderation Queue",
     "The admin moderation panel displays all user-reported content in a prioritised queue. Admins can review each flagged item in full context, view the report history, and take action: dismiss the report, issue a warning, delete content, or escalate to user suspension/ban.",
     "Admin","High",
     "1. Moderation queue shows all flagged content sorted by number of reports (highest first).\n2. Admin can view full post/comment content, number of reports, and report reasons.\n3. Admin actions: Dismiss Report, Delete Content, Warn User, Suspend User, Ban User.\n4. All actions are logged with admin ID, timestamp, reason, and action taken.\n5. Reporter is notified of the outcome of their report.\n6. Content with 5+ reports is automatically hidden until admin review.",
     "Phase 1","Milestone 1","FR-07-04, FR-10-02",
     "Moderation queue API with priority sorting; content_flags table; moderation_actions audit log; threshold-based auto-hide logic.",
     "Not Started"),

    # ── MODULE FR-11: Chairman Dashboard ──────────────────────────────────
    ("FR-11-01","Chairman Dashboard","Engagement Tracking",
     "The LGA Chairman's dashboard provides real-time engagement tracking for all content published by the LGA. Metrics tracked include: total views per post/project, comment counts, reaction breakdowns (Like, Dislike, Support, Question), and citizen feedback summaries. Data is presented in visual charts and sortable tables.",
     "LGA Chairman","High",
     "1. Dashboard shows total views, comments, and reactions for all LGA content.\n2. Per-post breakdown: views, unique visitors, reaction counts by type.\n3. Trend graphs show engagement over last 7 days, 30 days, and 3 months.\n4. Top 5 most-engaged posts are highlighted.\n5. Data updates at minimum every 30 minutes.\n6. Chairman can filter data by date range and project category.",
     "Phase 1","Milestone 1","FR-02-01, FR-06-01",
     "Analytics aggregation views in PostgreSQL; Dashboard charts via Recharts; server-side data fetching via Next.js API routes.",
     "Not Started"),

    ("FR-11-02","Chairman Dashboard","AI Sentiment Analysis",
     "The Chairman's dashboard includes an AI-powered sentiment analysis module that analyses all citizen comments and feedback on their LGA's content. The system categorises public sentiment as Positive, Neutral, or Negative, and generates readable summary reports showing what citizens are saying.",
     "LGA Chairman","Medium",
     "1. Sentiment analysis runs on all new comments posted on LGA content.\n2. Dashboard shows an overall sentiment score (% Positive, % Neutral, % Negative).\n3. Sentiment trend graph shows changes over the last 30 days.\n4. Top positive and top negative comment themes are extracted and shown.\n5. Sentiment is recalculated at least once per day.\n6. Chairman can view sentiment for individual posts.",
     "Phase 2","Milestone 2","FR-07-02",
     "OpenAI API or HuggingFace sentiment model; sentiment_scores table per comment; daily batch processing cron job.",
     "Not Started"),

    ("FR-11-03","Chairman Dashboard","Monthly Performance Reports",
     "The Chairman's dashboard generates comprehensive monthly performance reports summarising: total engagement metrics, project status overview, citizen registration growth, sentiment summary, and subscription status. Reports are automatically generated on the 1st of each month and are downloadable as PDF.",
     "LGA Chairman","Medium",
     "1. Monthly reports are auto-generated on the 1st of every month.\n2. Report covers the previous full calendar month's data.\n3. Sections include: Engagement Summary, Top Projects, Citizen Growth, Sentiment Overview, Subscription Status.\n4. Chairman receives an email when the report is ready with a download link.\n5. Last 12 monthly reports are stored and accessible on the dashboard.\n6. Admin can also access all Chairman reports from the admin panel.",
     "Phase 2","Milestone 2","FR-11-01, FR-11-02",
     "Monthly cron job on 1st of month; Puppeteer PDF generation; report stored on CDN; email notification via SMTP.",
     "Not Started"),

    ("FR-11-04","Chairman Dashboard","Post Scheduling",
     "The Chairman (and authorised staff) can schedule project updates and announcements to be automatically published at a specified future date and time. This allows Chairmen to plan their communications in advance and maintain a consistent publishing schedule.",
     "LGA Chairman, LGA Staff","Medium",
     "1. Post creation form includes a 'Schedule Publication' option with date/time picker.\n2. Scheduled posts are saved as 'Scheduled' status and are not visible to the public.\n3. A cron job checks for scheduled posts every minute and publishes them at the specified time.\n4. Chairman can view all scheduled posts in a calendar view on the dashboard.\n5. Scheduled posts can be edited or cancelled before their scheduled time.\n6. Published (previously scheduled) posts show the scheduled time, not the actual publish time.",
     "Phase 2","Milestone 2","FR-06-01",
     "scheduled_at field on posts table; cron job (every minute) to check and publish; calendar view component in dashboard.",
     "Not Started"),

    ("FR-11-05","Chairman Dashboard","Subscription Renewal Tracking",
     "The Chairman's dashboard prominently displays the current subscription status, remaining days until expiry, and provides direct subscription renewal functionality. Automated reminder emails are sent at 30 days, 14 days, 7 days, and daily during the grace period.",
     "LGA Chairman","High",
     "1. Dashboard header shows: plan type, expiry date, and days remaining (highlighted in amber/red when <30 days).\n2. 'Renew Subscription' button launches Paystack checkout directly from the dashboard.\n3. Renewal extends the subscription from the current expiry date (not from today).\n4. Reminder emails are sent at: 30 days, 14 days, 7 days, and every day during the 7-day grace period.\n5. Chairman can switch between Monthly and Annual plans on renewal.",
     "Phase 2","Milestone 2","FR-03-01, FR-03-04",
     "Cron job for daily reminder checks; subscription_expiry_date on LGA table; Paystack renewal checkout integration.",
     "Not Started"),

    ("FR-11-06","Chairman Dashboard","Citizen Registration Statistics",
     "The Chairman's dashboard displays the total number of citizens who have registered on the platform and selected their LGA as the Chairman's LGA. This metric helps Chairmen gauge their community's digital engagement and platform adoption.",
     "LGA Chairman","Medium",
     "1. Dashboard shows total citizen count for the LGA (real-time count from database).\n2. A trend graph shows monthly citizen registration growth for the LGA.\n3. Comparative indicator shows if the LGA's citizen count is above or below average for their State.\n4. Citizen registration data is anonymised (no personal data exposed to Chairman).\n5. Admin can view citizen counts for all LGAs in a sortable table.",
     "Phase 1","Milestone 1","FR-01-01, FR-02-01",
     "Aggregate COUNT query on users table filtered by lga_id; cached hourly; comparison computed against State average.",
     "Not Started"),

    # ── MODULE FR-12: Data Management & Archiving ──────────────────────────
    ("FR-12-01","Data Management & Archiving","Archived Content Accessibility",
     "All content published on the platform (posts, projects, allocation data, press releases) is permanently archived and remains publicly accessible after a Chairman's tenure ends or an LGA account is suspended. Archived content is clearly labelled with the relevant tenure period.",
     "All Users","High",
     "1. Archived content is accessible via the LGA's public profile page under an 'Archive' tab.\n2. Content is labelled with the Chairman's name and tenure period (e.g., 'Published under [Name], 2023-2025').\n3. Archived content cannot be edited or deleted by current or past Chairmen.\n4. Admin can hide (but not delete) archived content flagged for policy violations.\n5. Archived content is included in platform-wide search results.",
     "Phase 2","Milestone 2","FR-02-01, FR-06-01",
     "is_archived flag; tenure_id linking posts to specific chairman tenures; read-only access for archived content.",
     "Not Started"),

    ("FR-12-02","Data Management & Archiving","Re-election Data Restoration",
     "When a Chairman is re-elected and their re-election is verified by the admin, their full profile including all historical posts, projects, and engagement data from their previous tenure is restored and linked to their new tenure. A new tenure period is created but connected to all previous records.",
     "LGA Chairman, Admin","High",
     "1. Upon admin approval of re-election, system creates a new tenure record linked to the existing Chairman profile.\n2. All previous posts, projects, and engagement data are linked and accessible.\n3. New tenure start date is set to the date of admin approval.\n4. Chairman's dashboard shows both current tenure data and cumulative historical data.\n5. Public profile shows posts organised by tenure with clear labelling.",
     "Phase 2","Milestone 2","FR-03-03, FR-12-01",
     "chairman_tenures table with (chairman_id, lga_id, start_date, end_date); posts linked via tenure_id; multi-tenure query logic.",
     "Not Started"),

    ("FR-12-03","Data Management & Archiving","Engagement Data Retention",
     "All engagement data (Likes, Dislikes, Comments, Views) accumulated on any post or project is permanently retained even after the originating Chairman's tenure ends, the account is suspended, or the Chairman is re-elected. Engagement data is linked to the specific tenure for historical accuracy.",
     "All Users","Medium",
     "1. Engagement counts (likes, dislikes, comments, views) are not reset between tenures.\n2. Cumulative engagement totals are displayed on the public LGA page.\n3. Per-tenure engagement breakdown is accessible via the tenure archive view.\n4. Engagement data is included in admin analytics reports.\n5. Citizens' engagement history is preserved even if the LGA account is suspended.",
     "Phase 2","Milestone 2","FR-07-01, FR-07-02",
     "Reactions and comments tables permanently retain all records; tenure_id FK allows per-tenure filtering while cumulative queries ignore tenure.",
     "Not Started"),

    ("FR-12-04","Data Management & Archiving","Chairman Succession Record",
     "The platform maintains a complete, chronological succession record for each LGA, listing all Chairmen who have held the position and registered on the platform. When a new Chairman registers for an LGA, they inherit the LGA page but create a new tenure-specific record.",
     "All Users, Admin","Medium",
     "1. LGA public profile page has a 'Leadership History' section listing all Chairmen.\n2. Each entry shows: Chairman's name, tenure start/end dates, and key stats (projects, total allocation, engagement).\n3. Clicking a historical Chairman's entry navigates to their archived tenure page.\n4. New Chairman registration checks if the LGA has an existing record and links appropriately.\n5. Admin can merge or correct succession records if data errors occur.",
     "Phase 2","Milestone 2","FR-02-01, FR-12-01",
     "chairman_tenures table; LGA leadership history API; public tenure browsing component.",
     "Not Started"),

    ("FR-12-05","Data Management & Archiving","Public Historical Data Search",
     "Citizens can search and browse historical platform data by: specific tenure periods, Chairman names, and LGA name. Advanced search filters allow navigation through different administration periods, allocation data by year, and project history. All historical data is searchable without requiring login.",
     "Citizens","Medium",
     "1. A dedicated 'Historical Data' search page is available on the platform.\n2. Search accepts: LGA name, Chairman name, year range, project category, allocation period.\n3. Results are paginated and sorted by relevance or date.\n4. Search results include: posts, projects, allocation records, and press releases.\n5. No login is required to search historical data.\n6. Search returns results within 2 seconds for standard queries.",
     "Phase 2","Milestone 2","FR-12-01, FR-12-04",
     "PostgreSQL full-text search (tsvector) or Elasticsearch index; search API with multiple filter parameters; public route (no auth required).",
     "Not Started"),

    # ── MODULE FR-13: Government Collaboration & Transparency ──────────────
    ("FR-13-01","Government Collaboration & Transparency","Government Data Integration",
     "The platform integrates with Federal and State Government data systems to automatically pull and display budget allocation data per LGA. When government data APIs are available, allocation figures are ingested automatically. Where APIs are unavailable, admin manual data entry is the fallback.",
     "Admin, Citizens","Medium",
     "1. API integration with Federal Government's open data portal (where available).\n2. Scheduled data pull (daily or weekly) to fetch latest allocation data.\n3. Data is transformed and stored in the platform's allocation_records table.\n4. Admin is notified when new data is pulled successfully or when the pull fails.\n5. Manual data entry remains available as a fallback for admin.\n6. Data provenance (source, pull date) is displayed alongside published figures.",
     "Phase 2","Milestone 2","FR-04-01",
     "Government API integration via HTTP client (Axios); data transformation pipeline; scheduled cron job; fallback to admin manual entry.",
     "Not Started"),

    ("FR-13-02","Government Collaboration & Transparency","Procurement Portal Integration",
     "The platform integrates with the Bureau of Public Procurement (BPP) or equivalent government procurement portal to display awarded public contracts linked to each LGA. Citizens can view contractor names, contract values, project scope, and award dates for their LGA.",
     "Citizens, Admin","Low",
     "1. Procurement data is displayed on each LGA's profile page under a 'Public Contracts' tab.\n2. Each entry shows: contract title, contractor name, awarded value (NGN), award date, and project scope.\n3. Data is sourced from government procurement APIs or manually entered by admin.\n4. Citizens can flag a contract for review if they believe data is incorrect.\n5. Procurement data is searchable and filterable by value, date, and contractor name.",
     "Phase 2","Milestone 2","FR-13-01",
     "BPP API integration (where available); procurement_contracts table; public read-only display; contract_flags for citizen reporting.",
     "Not Started"),

    ("FR-13-03","Government Collaboration & Transparency","Audit Report Access",
     "Citizens can access publicly released audit reports for LGAs where available. The platform provides a structured repository of audit reports (uploaded by admin from government sources) that allows citizens to monitor fund utilisation and financial accountability at the LGA level.",
     "Citizens, Admin","Low",
     "1. Admin can upload PDF audit reports linked to specific LGAs and financial years.\n2. Uploaded reports appear on the LGA's profile under an 'Audit Reports' tab.\n3. Citizens can download any published audit report without login.\n4. Audit reports are labelled with: LGA, financial year, auditing body, and upload date.\n5. Citizens can leave comments on audit reports (subject to same moderation rules).",
     "Phase 2","Milestone 2","FR-10-04",
     "audit_reports table with lga_id, financial_year, report_url, uploaded_by, created_at; PDF stored on CDN.",
     "Not Started"),

    # ── MODULE FR-14: News & Press Integration ─────────────────────────────
    ("FR-14-01","News & Press Integration","Press Release Publishing",
     "Official government press releases from LGAs, State governments, and Federal agencies can be published on the platform by admin. Press releases are prominently featured in a dedicated 'News & Press' section and are linkable from LGA profiles.",
     "Admin, LGA Chairman","Medium",
     "1. Admin can create and publish press releases with a rich text editor.\n2. Press releases include: headline, body, issuing entity, date issued, and optional attachments.\n3. Press releases appear in the 'News & Press' section chronologically.\n4. Citizens can share press releases via WhatsApp, Twitter/X, and Facebook.\n5. Press releases are indexed for platform search.\n6. LGA Chairmen can also submit press releases that require admin approval before publication.",
     "Phase 2","Milestone 2","FR-10-04",
     "press_releases table; rich text editor in admin CMS and Chairman dashboard; submission workflow for Chairman-initiated press releases.",
     "Not Started"),

    ("FR-14-02","News & Press Integration","Live Streaming",
     "The platform supports live streaming of major government events including: project launch ceremonies, LGA town halls, budget presentations, and state government press briefings. Live streams are embedded directly on the platform using YouTube Live or a direct RTMP stream integration.",
     "Admin, LGA Chairman","Low",
     "1. Admin or authorised Chairman can schedule a live stream event from their dashboard.\n2. A scheduled live stream event appears as an upcoming event on the platform with countdown.\n3. Live stream is embedded via YouTube Live URL or direct RTMP/HLS stream.\n4. Citizens can watch the live stream and interact via a live comment sidebar.\n5. Live stream is automatically archived on the platform after it ends.",
     "Phase 2","Milestone 2","FR-14-01",
     "YouTube Live API for stream embedding; live_streams table with scheduled_at, stream_url, status; live comment integration.",
     "Not Started"),

    ("FR-14-03","News & Press Integration","Archived Press Releases",
     "All press releases published on the platform are permanently archived and remain publicly accessible. Citizens can browse the full archive of press releases, filter by issuing entity, date range, and LGA, and download individual press releases as PDF.",
     "Citizens","Medium",
     "1. A dedicated 'Press Archive' page lists all published press releases.\n2. Citizens can filter by: issuing entity (LGA/State/Federal), date range, and LGA.\n3. Search within the press archive is supported (full-text search).\n4. Each press release has a unique URL and is shareable.\n5. Press releases can be downloaded as PDF by any visitor (no login required).\n6. Archive is paginated with 20 entries per page.",
     "Phase 2","Milestone 2","FR-14-01",
     "Archived press releases are never deleted (is_published flag only); public-facing API route with filters; PDF export via Puppeteer.",
     "Not Started"),

    # ── MODULE FR-15: Additional Features ─────────────────────────────────
    ("FR-15-01","Additional Features & Future Expansion","Appointees for Posting",
     "LGA Chairmen can appoint up to two (2) individuals (outside of registered staff) to post content on the platform on their behalf. All content posted by appointees is subject to Chairman approval before it appears publicly. Appointees have the same limited permissions as registered staff.",
     "LGA Chairman","Medium",
     "1. Chairman can add up to 2 appointees via the dashboard (in addition to 2 staff: total 4 delegates).\n2. Appointee registration requires name, email, and phone.\n3. Appointees receive an invitation email to create a limited-access account.\n4. Appointee posts are held for Chairman approval before publication.\n5. Appointee access can be revoked by the Chairman at any time.\n6. Appointees cannot access payment, subscription, or account settings.",
     "Phase 2","Milestone 2","FR-02-05",
     "Appointees stored in same staff/delegates table with role='appointee'; invitation flow via email; same approval workflow as staff.",
     "Not Started"),

    ("FR-15-02","Additional Features & Future Expansion","Google Maps API Integration",
     "The platform integrates with the Google Maps API to display precise project locations on interactive maps embedded within project pages and the LGA profile. Project markers show popup cards with project title, status, category, and a media preview (first uploaded photo/video thumbnail).",
     "All Users","High",
     "1. Each project page shows an embedded Google Map with a pin at the project's exact location.\n2. Clicking the pin opens a popup with project name, status, and photo thumbnail.\n3. LGA profile page shows a mini-map with all the LGA's project pins.\n4. Project location is set by clicking on the map during project creation (geocoding).\n5. Google Maps API key is securely stored server-side and never exposed to the client.",
     "Phase 1","Milestone 1","FR-06-02",
     "Google Maps JavaScript API and Geocoding API; API key stored in server env vars; Next.js API proxy for map data to avoid key exposure.",
     "Not Started"),

    ("FR-15-03","Additional Features & Future Expansion","Multi-Language Support",
     "The platform supports five languages to serve Nigeria's linguistically diverse population: English (default), Hausa, Igbo, Yoruba, and Pidgin English. Users can switch languages from any page using a language selector in the navigation bar. UI text, labels, and error messages are fully translated.",
     "All Users","Medium",
     "1. Language selector is prominently placed in the navigation header.\n2. Selected language preference is saved to user's browser (localStorage) and account (if logged in).\n3. Switching language updates all UI text, labels, buttons, and form placeholders immediately.\n4. All 5 languages (English, Hausa, Igbo, Yoruba, Pidgin) are fully translated for all UI components.\n5. User-generated content (posts, comments) is not automatically translated.",
     "Phase 2","Milestone 2","None",
     "next-i18next or i18next for Next.js internationalisation; translation JSON files for each language; language preference stored in user profile.",
     "Not Started"),

    ("FR-15-04","Additional Features & Future Expansion","Accessibility Features",
     "The platform implements comprehensive accessibility features to ensure inclusivity for all citizens including visually impaired and differently-abled users. Features include: Voice-to-Text input, Text-to-Speech output, Dark Mode, High Contrast Mode, and full keyboard navigation support.",
     "All Users","Medium",
     "1. Voice-to-Text: Users can dictate text in comment and feedback fields (browser Web Speech API).\n2. Text-to-Speech: Any text content can be read aloud via a 'Listen' button.\n3. Dark Mode: Toggle available in navbar; preference saved to account/localStorage.\n4. High Contrast Mode: Meets WCAG AA contrast ratio standards (4.5:1 minimum).\n5. All interactive elements are accessible via keyboard (Tab, Enter, Space, Arrow keys).\n6. All images have descriptive alt text; all form fields have proper labels.",
     "Phase 2","Milestone 2","None",
     "Web Speech API for voice input/output; CSS custom properties for theme switching; WCAG 2.1 AA compliance target; axe-core for automated accessibility testing.",
     "Not Started"),

    ("FR-15-05","Additional Features & Future Expansion","Future Mobile App Features",
     "A future mobile application (iOS and Android) is planned as an extension of the web platform. Future mobile-exclusive features include: Push Notifications for government updates and project changes, Offline Mode for browsing previously viewed content, and QR Code Scanning for on-site project verification.",
     "Citizens, LGA Chairman","Low",
     "1. [Future] Push notifications sent when LGA posts new content, projects are updated, or allocation data is published.\n2. [Future] Offline Mode caches last 50 posts for browsing without internet connection.\n3. [Future] QR codes are generated for each project; scanning takes user directly to the project page.\n4. [Future] Mobile app uses the same API as the web platform.\n5. [Future] App complies with Apple App Store and Google Play Store guidelines.",
     "Future","Post-Launch","FR-06-01, FR-07-01",
     "React Native or Flutter for mobile; Firebase Cloud Messaging (FCM) for push notifications; QR code generation using qrcode.react; Service Workers for offline caching.",
     "Future Scope"),
]

# Write features to worksheet
for row_idx, feature in enumerate(features, start=4):
    module_prefix = feature[0][:5]
    colors = SECTION_COLORS.get(module_prefix, (HEADER_GRAY, WHITE))
    is_even = (row_idx % 2 == 0)
    row_bg = colors[1] if not is_even else WHITE

    for col_idx, value in enumerate(feature, start=1):
        cell = ws_fr.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
        cell.border = thin_border()
        cell.font = body_font(9)
        cell.fill = fill(row_bg)

        if col_idx == 1:  # Feature ID
            cell.font = Font(name="Arial", size=9, bold=True, color=colors[0])
            cell.alignment = center_align()
        elif col_idx == 2:  # Module
            cell.font = Font(name="Arial", size=9, bold=True, color=colors[0])
            cell.alignment = wrap_align("left", "top")
        elif col_idx == 3:  # Feature Name
            cell.font = Font(name="Arial", size=9, bold=True, color="212121")
            cell.alignment = wrap_align("left", "top")
        elif col_idx == 6:  # Priority
            priority = str(value)
            if priority == "High":
                cell.fill = fill(RED_LIGHT)
                cell.font = Font(name="Arial", size=9, bold=True, color=RED_DARK)
            elif priority == "Medium":
                cell.fill = fill(ORANGE_LIGHT)
                cell.font = Font(name="Arial", size=9, bold=True, color=ORANGE)
            else:
                cell.fill = fill(LIGHT_GREEN)
                cell.font = Font(name="Arial", size=9, bold=True, color=MID_GREEN)
            cell.alignment = center_align()
        elif col_idx == 8:  # Phase
            cell.alignment = center_align()
        elif col_idx == 9:  # Milestone
            cell.alignment = center_align()
        elif col_idx == 12:  # Status
            status = str(value)
            if status == "Not Started":
                cell.fill = fill("E3F2FD")
                cell.font = Font(name="Arial", size=9, bold=True, color=BLUE_DARK)
            elif status == "Future Scope":
                cell.fill = fill(PURPLE_LIGHT)
                cell.font = Font(name="Arial", size=9, bold=True, color=PURPLE)
            cell.alignment = center_align()
        else:
            cell.alignment = wrap_align("left", "top")

    ws_fr.row_dimensions[row_idx].height = 80

ws_fr.auto_filter.ref = f"A3:L{3 + len(features)}"

# ─── SHEET 3: NON-FUNCTIONAL REQUIREMENTS ─────────────────────────────────────
ws_nfr = wb.create_sheet("Non-Functional Requirements")
ws_nfr.sheet_view.showGridLines = False

ws_nfr.merge_cells("A1:J1")
ws_nfr["A1"].value = "NON-FUNCTIONAL REQUIREMENTS – LGA CITIZEN PORTAL"
ws_nfr["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_nfr["A1"].fill = fill(DARK_GREEN)
ws_nfr["A1"].alignment = center_align()
ws_nfr.row_dimensions[1].height = 30

ws_nfr.merge_cells("A2:J2")
ws_nfr["A2"].value = "System quality attributes defining how the platform should perform, scale, and maintain compliance"
ws_nfr["A2"].font = Font(name="Arial", size=10, italic=True, color=WHITE)
ws_nfr["A2"].fill = fill(MID_GREEN)
ws_nfr["A2"].alignment = center_align()
ws_nfr.row_dimensions[2].height = 18

nfr_headers = ["NFR ID", "Category", "Requirement", "Description", "Target Metric", "Priority", "Test Approach", "Phase", "Notes"]
nfr_widths   = [12, 16, 26, 48, 28, 10, 40, 12, 28]

for col_idx, (h, w) in enumerate(zip(nfr_headers, nfr_widths), start=1):
    cell = ws_nfr.cell(row=3, column=col_idx, value=h)
    cell.font = header_font(10, color=WHITE)
    cell.fill = fill(HEADER_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()
    ws_nfr.column_dimensions[get_column_letter(col_idx)].width = w
ws_nfr.row_dimensions[3].height = 28
ws_nfr.freeze_panes = "A4"

nfr_data = [
    # Performance
    ("NFR-01-01","Performance","Concurrent User Capacity",
     "The platform must handle up to 10,000 simultaneous active users across all user types (citizens, LGAs, admins, advertisers) without experiencing performance degradation, increased error rates, or unacceptable response times.",
     "10,000 concurrent users; <4s response time maintained","High",
     "Load testing with Apache JMeter or k6; simulate 10,000 concurrent virtual users; monitor response times and error rates.","Phase 3","Measured under realistic usage patterns (mix of read/write operations)."),

    ("NFR-01-02","Performance","Interactive Map Load Time",
     "The interactive Nigeria map showing all 774 LGA markers and boundaries must fully load and be interactive within 3 seconds on a standard 4G mobile connection (20 Mbps download speed).",
     "Map fully interactive within 3 seconds on 4G","High",
     "WebPageTest and Lighthouse testing on mobile device at simulated 4G speed; repeat 5 times and average.","Phase 3","Requires tile caching, GeoJSON optimisation, and CDN for tile delivery."),

    ("NFR-01-03","Performance","Page Load Time",
     "All platform pages (landing page, LGA profiles, project pages, dashboards, admin panel) must achieve a full page load time of under 4 seconds on a standard 4G connection. Core Web Vitals (LCP, FID, CLS) must meet Google's 'Good' thresholds.",
     "< 4 seconds on 4G; LCP < 2.5s; CLS < 0.1; FID < 100ms","High",
     "Google Lighthouse audit (target score 85+); WebPageTest field data; Real User Monitoring (RUM) post-launch.","Phase 3","Next.js SSR and static generation; image optimisation via Next/Image; CDN for static assets."),

    # Security
    ("NFR-02-01","Security","Data Encryption",
     "All user data, LGA data, payment data, and system communications must be encrypted both in transit (HTTPS/TLS 1.2+) and at rest (AES-256 encryption for database storage). Sensitive fields (passwords, OTPs) must be additionally hashed (bcrypt).",
     "TLS 1.2+ for transit; AES-256 at rest; bcrypt for passwords","High",
     "SSL Labs rating A or better; database encryption verification; penetration testing for data exposure vulnerabilities.","Phase 3","Passwords stored as bcrypt hashes (cost factor 12+); OTPs are time-limited and single-use."),

    ("NFR-02-02","Security","JWT Authentication",
     "All authenticated user sessions must use JSON Web Tokens (JWT) for stateless, secure session management. JWTs must be short-lived (access token: 15 minutes; refresh token: 7 days) and stored securely (httpOnly cookies; not localStorage).",
     "Access token: 15 min TTL; Refresh token: 7 days; httpOnly cookies","High",
     "Security audit of token implementation; OWASP JWT testing checklist; penetration test for token replay attacks.","Phase 1","NextAuth.js or custom JWT middleware in Next.js; token rotation on refresh; revocation list for critical actions."),

    ("NFR-02-03","Security","Security Audits & NDPR Compliance",
     "The platform must undergo regular security audits (minimum quarterly) to ensure compliance with the Nigeria Data Protection Regulation (NDPR) and other applicable data protection laws. A Data Protection Impact Assessment (DPIA) must be completed before launch.",
     "Quarterly security audits; NDPR compliance certificate pre-launch","High",
     "Third-party penetration testing pre-launch; NDPR compliance checklist review; automated vulnerability scanning (OWASP ZAP) in CI/CD.","Phase 3","Privacy policy, cookie policy, and data processing agreements must be in place before public launch."),

    ("NFR-02-04","Security","Payment Security (PCI-DSS)",
     "All payment processing must comply with Payment Card Industry Data Security Standard (PCI-DSS). The platform must not store, process, or transmit raw card data. All payment data is handled exclusively by Paystack, which is PCI-DSS Level 1 certified.",
     "PCI-DSS compliant; no raw card data stored on platform servers","High",
     "Paystack integration review; verify no card data stored in logs or database; Paystack webhook signature verification.","Phase 2","Platform is 'SAQ A' merchant type (card data handled entirely by Paystack); webhook secret key must be env variable."),

    # Scalability
    ("NFR-03-01","Scalability","Horizontal Scalability",
     "The platform architecture must support horizontal scaling to accommodate future growth beyond 774 LGAs and millions of citizens. The backend must be stateless (JWT sessions) and deployable across multiple server instances behind a load balancer.",
     "Support 3x traffic growth without architectural changes","Medium",
     "Load testing at 3x expected capacity; verify stateless session handling; test load balancer configuration.","Phase 3","Cloud hosting (AWS/DigitalOcean); container-based deployment (Docker + Kubernetes or Docker Compose); stateless Next.js API routes."),

    ("NFR-03-02","Scalability","Database Scalability",
     "The PostgreSQL database must be configured to efficiently handle large volumes of data including: millions of citizen records, hundreds of thousands of project updates, multi-year allocation records, media metadata, and engagement data (comments, reactions, views).",
     "Support 10M+ records; query response < 500ms for indexed queries","High",
     "Database load testing with production-scale data volumes; explain/analyze queries; monitor query performance with pg_stat_statements.","Phase 3","Read replicas for analytics queries; proper indexing strategy; PostGIS optimisation for spatial queries; partitioning for large tables."),

    # Usability
    ("NFR-04-01","Usability","Intuitive Navigation",
     "The platform must be designed to be intuitive and easy to navigate for all user types including low-digital-literacy citizens and first-time users. Navigation must follow established UX conventions with clear information hierarchy, consistent layouts, and discoverable features.",
     "System Usability Scale (SUS) score ≥ 70 in usability testing","High",
     "Usability testing with 10+ representative users from target demographic; SUS questionnaire; task completion rate > 85%.","Phase 1","User testing with both urban and rural demographics; simplified onboarding flow; contextual help tooltips."),

    ("NFR-04-02","Usability","Responsive Design",
     "The platform must provide an optimal, fully functional experience across all device types and screen sizes: desktop (1920px+), laptop (1280px–1920px), tablet (768px–1280px), and mobile (320px–768px). No horizontal scrolling on any device. All features accessible on mobile.",
     "Zero horizontal scroll on any device; 100% feature parity on mobile","High",
     "Cross-browser testing (Chrome, Firefox, Safari, Edge); device testing (iPhone, Android, iPad, desktop); BrowserStack or LambdaTest.","Phase 3","Mobile-first CSS design; Tailwind CSS responsive utilities; Next.js responsive image component."),

    ("NFR-04-03","Usability","Accessibility (WCAG 2.1)",
     "The platform must meet WCAG 2.1 Level AA accessibility standards to ensure the platform is usable by citizens with disabilities. This includes sufficient colour contrast, screen reader compatibility, keyboard navigability, and support for the platform's built-in accessibility features (Dark Mode, High Contrast, Voice features).",
     "WCAG 2.1 AA compliance; 0 critical accessibility errors in axe-core audit","Medium",
     "Automated axe-core testing in CI/CD pipeline; manual screen reader testing (NVDA/VoiceOver); keyboard-only navigation audit.","Phase 3","Use semantic HTML5 elements; ARIA labels where needed; focus management for modals/dialogs; skip-to-content links."),
]

for row_idx, nfr in enumerate(nfr_data, start=4):
    cat = nfr[1]
    cat_colors = {
        "Performance": (MID_GREEN, LIGHT_GREEN),
        "Security": (RED_DARK, RED_LIGHT),
        "Scalability": (BLUE_DARK, BLUE_LIGHT),
        "Usability": (ORANGE, ORANGE_LIGHT),
    }
    c_dark, c_light = cat_colors.get(cat, (HEADER_GRAY, WHITE))
    row_bg = c_light if row_idx % 2 == 0 else WHITE

    for col_idx, value in enumerate(nfr, start=1):
        cell = ws_nfr.cell(row=row_idx, column=col_idx, value=str(value))
        cell.border = thin_border()
        cell.font = body_font(9)
        cell.fill = fill(row_bg)

        if col_idx == 1:
            cell.font = Font(name="Arial", size=9, bold=True, color=c_dark)
            cell.alignment = center_align()
        elif col_idx == 2:
            cell.font = Font(name="Arial", size=9, bold=True, color=c_dark)
            cell.alignment = center_align()
        elif col_idx == 3:
            cell.font = Font(name="Arial", size=9, bold=True, color="212121")
            cell.alignment = wrap_align("left", "top")
        elif col_idx == 6:
            cell.font = Font(name="Arial", size=9, bold=True, color=RED_DARK)
            cell.fill = fill(RED_LIGHT)
            cell.alignment = center_align()
        elif col_idx == 8:
            cell.alignment = center_align()
        else:
            cell.alignment = wrap_align("left", "top")

    ws_nfr.row_dimensions[row_idx].height = 70

ws_nfr.auto_filter.ref = f"A3:I{3 + len(nfr_data)}"

# ─── SHEET 4: PHASE MAPPING ───────────────────────────────────────────────────
ws_phase = wb.create_sheet("Phase & Milestone Mapping")
ws_phase.sheet_view.showGridLines = False

ws_phase.merge_cells("A1:H1")
ws_phase["A1"].value = "DEVELOPMENT PHASE & MILESTONE MAPPING"
ws_phase["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_phase["A1"].fill = fill(DARK_GREEN)
ws_phase["A1"].alignment = center_align()
ws_phase.row_dimensions[1].height = 30

phases = [
    ("Phase 1", "Requirements, Design & Core Development", "5 Weeks", "Milestone 1 – 40% (NGN 384,000)",
     [("FR-01-01","Citizen Registration"),("FR-01-02","Citizen Dashboard"),("FR-01-03","Engagement Tools"),
      ("FR-01-04","LGA Citizen Statistics"),("FR-01-05","OTP Authentication"),
      ("FR-02-01","LGA Profile Creation"),("FR-02-02","Verification Process"),("FR-02-03","Admin Approval"),
      ("FR-02-05","Staff Registration"),("FR-05-01","Interactive Map"),("FR-05-02","LGA Detail Pop-up"),
      ("FR-05-03","Colour-Coded Markers"),("FR-05-04","Project Status Tracking"),("FR-05-05","Archived Projects on Map"),
      ("FR-06-01","Project Publication"),("FR-06-02","Project Detail Requirements"),
      ("FR-06-03","Verification Prerequisite"),("FR-06-04","Rich Media Support"),
      ("FR-07-01","Reaction System"),("FR-07-02","Comments & Feedback"),("FR-07-03","Social Media Sharing"),
      ("FR-07-04","Content Reporting System"),("FR-07-06","Admin Moderation Authority"),
      ("FR-10-06","Content Moderation Queue"),("FR-11-01","Engagement Tracking"),
      ("FR-11-06","Citizen Registration Stats"),("FR-15-02","Google Maps API Integration")]),

    ("Phase 2", "Advanced Development", "5 Weeks", "Milestone 2 – 35% (NGN 336,000)",
     [("FR-02-04","Admin Charges & Free Period"),("FR-03-01","LGA Subscription Fee"),
      ("FR-03-02","Tenure Management"),("FR-03-03","Re-election Process"),("FR-03-04","Grace Period"),
      ("FR-03-05","Staff Posting Rights Control"),("FR-04-01","Monthly Allocation Data"),
      ("FR-04-02","Admin Content Creation"),("FR-04-03","Historical Allocation Archive"),
      ("FR-04-04","Allocation Comparison Tool"),("FR-07-05","AI Content Moderation"),
      ("FR-08-01","Ads Placement System"),("FR-08-02","Ad Subscription Plans"),("FR-08-03","Ad Formats"),
      ("FR-08-04","Ad Payment via Paystack"),("FR-08-05","Advertiser Dashboard"),("FR-08-06","Admin Ad Management"),
      ("FR-09-01","Paystack Payment Gateway"),("FR-09-02","Email Notifications"),
      ("FR-09-03","Payment History"),("FR-09-04","Invoice Generation"),
      ("FR-10-01","Full LGA Management"),("FR-10-02","User Account Management"),
      ("FR-10-03","Ads Placement Management"),("FR-10-04","Content Creation & Publishing"),
      ("FR-10-05","Engagement Analytics Dashboard"),("FR-11-02","AI Sentiment Analysis"),
      ("FR-11-03","Monthly Performance Reports"),("FR-11-04","Post Scheduling"),
      ("FR-11-05","Subscription Renewal Tracking"),("FR-12-01","Archived Content Accessibility"),
      ("FR-12-02","Re-election Data Restoration"),("FR-12-03","Engagement Data Retention"),
      ("FR-12-04","Chairman Succession Record"),("FR-12-05","Public Historical Data Search"),
      ("FR-13-01","Government Data Integration"),("FR-13-02","Procurement Portal Integration"),
      ("FR-13-03","Audit Report Access"),("FR-14-01","Press Release Publishing"),
      ("FR-14-02","Live Streaming"),("FR-14-03","Archived Press Releases"),
      ("FR-15-01","Appointees for Posting"),("FR-15-03","Multi-Language Support"),
      ("FR-15-04","Accessibility Features")]),

    ("Phase 3", "Testing, QA & Deployment", "4 Weeks", "Milestone 3 – 25% (NGN 240,000)",
     [("UAT","User Acceptance Testing"),("SEC-TEST","Security & Penetration Testing"),
      ("PERF-TEST","Performance & Load Testing"),("DEPLOY","Production Deployment"),
      ("SOFT-LAUNCH","Soft Launch with Select LGAs"),("NFR-01-01","Concurrent User Capacity Test"),
      ("NFR-01-02","Map Load Time Test"),("NFR-02-01","Data Encryption Verification"),
      ("NFR-04-02","Responsive Design Testing"),("NFR-04-03","WCAG 2.1 Audit")]),

    ("Phase 4", "Post-Launch Support & Maintenance", "Ongoing", "Milestone 4 – 7%/month (NGN 67,200/month)",
     [("MONITOR","Platform Performance Monitoring"),("BUG-FIX","Bug Fixes & Patches"),
      ("SECURITY-UPD","Security Updates"),("PERF-OPT","Performance Optimisation"),
      ("FEATURE-ENH","Minor Feature Enhancements"),("FR-15-05","Mobile App Features (Future Scope)")]),
]

p_colors = [MID_GREEN, BLUE_DARK, ORANGE, PURPLE]

col_letters_phase = ["A","B","C","D","E","F","G","H"]
phase_col_widths   = [12, 32, 12, 36, 14, 28, 14, 16]
phase_headers = ["Phase", "Phase Name", "Duration", "Milestone & Payment", "Feature ID", "Feature Name", "PRD Section", "Priority"]
r = 3
for col_idx, (h, w) in enumerate(zip(phase_headers, phase_col_widths), start=1):
    cell = ws_phase.cell(row=r, column=col_idx, value=h)
    cell.font = header_font(10, color=WHITE)
    cell.fill = fill(HEADER_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()
    ws_phase.column_dimensions[get_column_letter(col_idx)].width = w
ws_phase.row_dimensions[r].height = 28
ws_phase.freeze_panes = "A4"

r = 4
for p_idx, (pname, pdesc, pdur, pmile, pfeatures) in enumerate(phases):
    p_dark = p_colors[p_idx]
    p_start = r
    for f_idx, (fid, fname) in enumerate(pfeatures):
        row_bg = "F8F9FA" if f_idx % 2 == 0 else WHITE
        ws_phase.cell(row=r, column=1, value=pname).border = thin_border()
        ws_phase.cell(row=r, column=2, value=pdesc).border = thin_border()
        ws_phase.cell(row=r, column=3, value=pdur).border = thin_border()
        ws_phase.cell(row=r, column=4, value=pmile).border = thin_border()
        ws_phase.cell(row=r, column=5, value=fid).border = thin_border()
        ws_phase.cell(row=r, column=6, value=fname).border = thin_border()
        ws_phase.cell(row=r, column=7, value="FR-"+fid.split("-")[1] if fid.startswith("FR") else "N/A").border = thin_border()
        ws_phase.cell(row=r, column=8, value="High" if p_idx < 2 else "Medium").border = thin_border()

        for c in range(1, 9):
            cell = ws_phase.cell(row=r, column=c)
            cell.fill = fill(row_bg)
            cell.font = body_font(9)
            cell.alignment = wrap_align("left", "center") if c not in [1,3,5,7,8] else center_align()
            if c in [1,2,3,4]:
                cell.font = Font(name="Arial", size=9, bold=(c <= 2), color=p_dark if c <= 2 else "212121")
        ws_phase.row_dimensions[r].height = 18
        r += 1

    if p_start < r - 1:
        for c in [1,2,3,4]:
            ws_phase.merge_cells(start_row=p_start, start_column=c, end_row=r-1, end_column=c)
            cell = ws_phase.cell(row=p_start, column=c)
            cell.font = Font(name="Arial", size=10, bold=True, color=WHITE if c <= 2 else "212121")
            cell.fill = fill(p_dark if c <= 2 else "ECEFF1")
            cell.alignment = center_align()
            cell.border = thin_border()

ws_phase.auto_filter.ref = f"E3:H{r-1}"

# ─── SHEET 5: USER ROLES MATRIX ───────────────────────────────────────────────
ws_roles = wb.create_sheet("User Roles & Permissions")
ws_roles.sheet_view.showGridLines = False

ws_roles.merge_cells("A1:G1")
ws_roles["A1"].value = "USER ROLES & FEATURE PERMISSIONS MATRIX"
ws_roles["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_roles["A1"].fill = fill(DARK_GREEN)
ws_roles["A1"].alignment = center_align()
ws_roles.row_dimensions[1].height = 30

ws_roles.merge_cells("A2:G2")
ws_roles["A2"].value = "Legend:  R = Read  |  C = Create  |  E = Edit  |  D = Delete  |  A = Approve  |  — = No Access"
ws_roles["A2"].font = Font(name="Arial", size=10, italic=True, color=WHITE)
ws_roles["A2"].fill = fill(MID_GREEN)
ws_roles["A2"].alignment = center_align()
ws_roles.row_dimensions[2].height = 18

role_headers = ["Feature Area", "Citizen", "LGA Chairman", "LGA Staff", "Advertiser", "Admin", "Notes"]
role_widths   = [36, 14, 16, 14, 14, 14, 32]

for col_idx, (h, w) in enumerate(zip(role_headers, role_widths), start=1):
    cell = ws_roles.cell(row=3, column=col_idx, value=h)
    cell.font = header_font(10, color=WHITE)
    cell.fill = fill(HEADER_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()
    ws_roles.column_dimensions[get_column_letter(col_idx)].width = w
ws_roles.row_dimensions[3].height = 28
ws_roles.freeze_panes = "A4"

roles_data = [
    ("Citizen Registration & Login",          "C, R, E",    "—",       "—",     "—",         "R, D",    "Citizens manage own accounts; Admin can suspend/ban"),
    ("Citizen Dashboard",                     "R",          "—",       "—",     "—",         "R",       "Citizens view personalised LGA feed"),
    ("LGA Profile Registration",              "—",          "C",       "—",     "—",         "R, A, D", "Chairman creates; Admin approves/rejects"),
    ("LGA Profile Management",                "—",          "R, E",    "R",     "—",         "R, E, D", "Chairman edits; Staff read-only; Admin full control"),
    ("LGA Verification Documents",            "—",          "C, R",    "—",     "—",         "R, A",    "Chairman uploads; Admin reviews and approves"),
    ("Staff Account Management",              "—",          "C, E, D", "R",     "—",         "R, D",    "Chairman manages staff; Admin can override"),
    ("Project Creation & Publishing",         "—",          "C, E, D", "C",     "—",         "R, D",    "Staff create (Chairman approval needed); Chairman publishes"),
    ("Project Approval (Staff Posts)",        "—",          "A",       "—",     "—",         "R",       "Chairman approves all staff-created content"),
    ("Post Reactions (Like/Dislike etc.)",    "C",          "—",       "—",     "—",         "D",       "Only citizens can react; Admin can remove reactions"),
    ("Comments & Feedback",                   "C, E, D",    "C, R",    "—",     "—",         "R, D",    "Citizens post; Chairman can reply; Admin moderates"),
    ("Content Reporting",                     "C",          "—",       "—",     "—",         "R, A, D", "Citizens report; Admin reviews and acts"),
    ("Social Sharing",                        "C",          "C",       "C",     "—",         "—",       "All registered users can share"),
    ("Allocation Data – Viewing",             "R",          "R",       "R",     "—",         "R, C, E", "Admin publishes; all roles can read"),
    ("Allocation Comparison Tool",            "R",          "R",       "—",     "—",         "R",       "Citizens and LGA can compare allocations"),
    ("Historical Data Search",               "R",           "R",       "R",     "—",         "R",       "Publicly accessible (no login required for viewing)"),
    ("Interactive Map",                       "R",          "R",       "R",     "R",         "R",       "Publicly viewable map; no login required"),
    ("Advertiser Account Registration",       "—",          "—",       "—",     "C, R, E",   "R, D",    "Advertisers manage own accounts"),
    ("Ad Campaign Creation",                  "—",          "—",       "—",     "C, R, E",   "R, A, D", "Advertisers create; Admin approves before display"),
    ("Ad Performance Dashboard",              "—",          "—",       "—",     "R",         "R",       "Advertisers view own campaign data; Admin views all"),
    ("Ad Plan Management",                    "—",          "—",       "—",     "R",         "C, E, D", "Admin creates/manages plans; Advertisers select"),
    ("Payment & Subscriptions",               "—",          "C, R",    "—",     "C, R",      "R",       "Chairman and Advertisers make payments via Paystack"),
    ("Invoice Download",                      "—",          "R",       "—",     "R",         "R",       "All payers can download their own invoices"),
    ("Chairman Analytics Dashboard",          "—",          "R",       "—",     "—",         "R",       "Chairman views LGA analytics; Admin views all"),
    ("AI Sentiment Analysis",                 "—",          "R",       "—",     "—",         "R",       "Generated by system; viewed by Chairman and Admin"),
    ("Post Scheduling",                       "—",          "C, E, D", "C",     "—",         "R",       "Chairman and Staff schedule; Chairman approves"),
    ("Subscription Management",               "—",          "C, E",    "—",     "C, E",      "R, E",    "Users manage own subscriptions; Admin can override"),
    ("Admin Content Publishing",              "—",          "—",       "—",     "—",         "C, R, E, D","Only Admin can publish allocation data and LGA news"),
    ("Press Releases",                        "R",          "C",       "—",     "—",         "C, A, D", "Admin publishes; Chairman submits for approval"),
    ("Live Streaming",                        "R",          "C",       "—",     "—",         "C, A",    "Chairman initiates; Admin approves and can end stream"),
    ("User Account Management",               "—",          "—",       "—",     "—",         "R, E, D", "Admin-only: suspend, ban, reinstate users"),
    ("LGA Approval/Suspension",               "—",          "—",       "—",     "—",         "R, A, D", "Admin-only: approve, reject, suspend, deactivate LGAs"),
    ("Platform-Wide Analytics",               "—",          "—",       "—",     "—",         "R",       "Admin-only: full platform engagement metrics"),
    ("Moderation Queue Management",           "—",          "—",       "—",     "—",         "R, A, D", "Admin-only: review flagged content and take action"),
    ("Audit Reports",                         "R",          "R",       "—",     "—",         "C, R, E", "Admin uploads; all can read"),
    ("Procurement Portal Data",               "R",          "R",       "—",     "—",         "R, C",    "Admin imports; citizens and LGAs view"),
    ("Chairman Succession Records",           "R",          "R",       "—",     "—",         "R, E",    "Publicly visible; Admin can edit records"),
]

for row_idx, row_data in enumerate(roles_data, start=4):
    alt = row_idx % 2 == 0
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws_roles.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border()
        cell.font = body_font(9, bold=(col_idx == 1))
        cell.fill = fill(ROW_ALT if alt else WHITE)
        if col_idx == 1:
            cell.alignment = wrap_align("left", "center")
        elif col_idx == 7:
            cell.alignment = wrap_align("left", "center")
            cell.font = Font(name="Arial", size=9, italic=True, color="546E7A")
        else:
            cell.alignment = center_align()
            # Color code access levels
            if val == "—":
                cell.font = Font(name="Arial", size=9, color="BDBDBD")
            elif "D" in val or "A" in val:
                cell.font = Font(name="Arial", size=9, bold=True, color=RED_DARK)
            elif "C" in val:
                cell.font = Font(name="Arial", size=9, bold=True, color=MID_GREEN)
            else:
                cell.font = Font(name="Arial", size=9, color=BLUE_DARK)
    ws_roles.row_dimensions[row_idx].height = 22

ws_roles.auto_filter.ref = f"A3:G{3 + len(roles_data)}"

# ─── FINAL SAVE ───────────────────────────────────────────────────────────────
output_path = r"C:\lgaportal\LGA_Portal_FRD_v1.0.xlsx"
wb.save(output_path)
print(f"FRD saved to: {output_path}")
